import os, io, json, psycopg2, psycopg2.extras, smtplib, calendar
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication
from flask import Flask, render_template, request, redirect, url_for, flash, g, jsonify, abort, Response
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import date, datetime, timedelta
from functools import wraps

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'ciclorh-dev-troque-em-producao')

login_manager = LoginManager(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Faça login para acessar o sistema.'
login_manager.login_message_category = 'warning'

DATABASE_URL = os.environ.get('DATABASE_URL', '')


class _Agente:
    nome = 'CICLO RH – apoio administrativo LTDA'
    cnpj = '32.075.028/0001-84'
    endereco = 'Av. Juracy Magalhães, 3340, Bloco A, sala 1104, Felícia'
    cidade = 'Vitória da Conquista – Bahia'
    email = 'adm.salmocosta@gmail.com'
    representante = 'Ildeflávio dos Santos Silva Maia'
    cargo = 'Sócio Gerente'
    cpf = '915.569.295-87'


AGENTE = _Agente()

ETAPAS_CRM = [
    'Lead Captado', 'Primeiro Contato', 'Qualificação',
    'Apresentação', 'Negociação', 'Contrato Enviado',
    'Contrato Assinado', 'Cliente Ativo'
]
ORIGENS_CRM = ['Site', 'Indicação', 'LinkedIn', 'Prospecção ativa', 'Eventos']
TIPOS_INTERACAO = ['Ligação', 'WhatsApp', 'E-mail', 'Reunião', 'Nota interna']
ITENS_IMPLANTACAO_PADRAO = [
    'Reunião de kick-off realizada',
    'TCE/Contrato assinado pelas partes',
    'Empresa cadastrada no sistema',
    'Supervisor cadastrado',
    'Áreas de estágio configuradas',
    'Vagas publicadas',
    'Primeiro estagiário alocado',
    'Follow-up 30 dias',
]
ETAPAS_CRM_COR = {
    'Lead Captado': '#6c757d',
    'Primeiro Contato': '#17a2b8',
    'Qualificação': '#0d6efd',
    'Apresentação': '#ffc107',
    'Negociação': '#fd7e14',
    'Contrato Enviado': '#6610f2',
    'Contrato Assinado': '#6f42c1',
    'Cliente Ativo': '#198754',
}

SEGMENTOS_PROSPECTO = [
    'Saúde', 'Educação', 'Comércio', 'Indústria', 'Tecnologia',
    'Serviços', 'Construção Civil', 'Alimentação', 'Agronegócio',
    'Jurídico / Contábil', 'Setor Público', 'Financeiro', 'Logística', 'Outro',
]
PORTES_EMPRESA = ['MEI', 'ME', 'EPP', 'Médio', 'Grande']
STATUS_PROSPECTO = ['novo', 'contatado', 'convertido', 'descartado']
CNAE_GRUPOS = [
    ('85 — Educação', '85'),
    ('86 — Atividades de atenção à saúde humana', '86'),
    ('87 — Atividades de atenção residencial', '87'),
    ('47 — Comércio varejista', '47'),
    ('46 — Comércio atacadista', '46'),
    ('45 — Comércio e reparação de veículos', '45'),
    ('62 — Tecnologia da informação', '62'),
    ('63 — Serviços de informação', '63'),
    ('41 — Construção de edifícios', '41'),
    ('42 — Obras de infraestrutura', '42'),
    ('43 — Serviços especializados de construção', '43'),
    ('56 — Alimentação e bebidas', '56'),
    ('10 — Fabricação de alimentos', '10'),
    ('64 — Serviços financeiros', '64'),
    ('65 — Seguros e previdência', '65'),
    ('69 — Jurídico e contabilidade', '69'),
    ('84 — Administração pública', '84'),
    ('01 — Agricultura e pecuária', '01'),
    ('49 — Transporte terrestre / logística', '49'),
    ('55 — Alojamento', '55'),
    ('71 — Arquitetura e engenharia', '71'),
    ('73 — Publicidade e pesquisa de mercado', '73'),
]
BAIRROS_VDC = sorted([
    'Alto Maron', 'Ângela Maron', 'Antônio Ferreira', 'Bateias',
    'Brasil', 'Candeias', 'Capinal', 'Centro', 'Conquista', 'Cruzeiro',
    'Espírito Santo', 'Felícia', 'Guarani', 'Ibirapuera', 'Índio Paraná',
    'Jurema', 'Lagoa das Flores', 'Leopoldina', 'Liberdade', 'Malhado',
    'Miro Cairo', 'Morada das Árvores', 'Nações', 'Nova Conquista',
    'Patagônia', 'Pedrinhas', 'Pradoso', 'Primavera', 'Recreio',
    'Renato Gonçalves', 'Santa Helena', 'Santa Inês', 'Santa Rita',
    'São Cristóvão', 'São Pedro', 'Sesqui', 'Simões Filho',
    'Universidade', 'Urbis I', 'Urbis II', 'Urbis III', 'Urbis IV',
    'Boa Vista', 'Zabelê',
])
SEGMENTO_COR = {
    'Saúde': '#059669', 'Educação': '#2563EB', 'Comércio': '#D97706',
    'Indústria': '#7C3AED', 'Tecnologia': '#0891B2', 'Serviços': '#6B7280',
    'Construção Civil': '#B45309', 'Alimentação': '#DC2626', 'Agronegócio': '#65A30D',
    'Jurídico / Contábil': '#4338CA', 'Setor Público': '#0369A1',
    'Financeiro': '#0F766E', 'Logística': '#C2410C', 'Outro': '#9CA3AF',
}


class User(UserMixin):
    def __init__(self, id, username, nome, role, acesso_crm=False, crm_role='vendedor'):
        self.id = str(id)
        self.username = username
        self.nome = nome
        self.role = role
        self.acesso_crm = bool(acesso_crm)
        self.crm_role = crm_role or 'vendedor'

    @property
    def is_admin(self):
        return self.role == 'admin'

    @property
    def is_crm(self):
        return self.is_admin or self.acesso_crm


@login_manager.user_loader
def load_user(user_id):
    row = _q("SELECT * FROM usuario WHERE id = %s", (user_id,), one=True)
    if row:
        return User(row['id'], row['username'], row['nome'], row['role'],
                    row.get('acesso_crm', False), row.get('crm_role', 'vendedor'))
    return None


def admin_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_admin:
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ─── BANCO DE DADOS ───────────────────────────────────────────────────────────

def _get_conn():
    if 'db' not in g:
        url = DATABASE_URL
        if url.startswith('postgres://'):
            url = url.replace('postgres://', 'postgresql://', 1)
        conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
        conn.autocommit = True
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(e=None):
    db = g.pop('db', None)
    if db and not db.closed:
        db.close()


def _q(sql, params=(), one=False):
    with _get_conn().cursor() as cur:
        cur.execute(sql, params)
        return cur.fetchone() if one else cur.fetchall()


def _run(sql, params=()):
    with _get_conn().cursor() as cur:
        cur.execute(sql, params)


def _ins(sql, params=()):
    with _get_conn().cursor() as cur:
        cur.execute(sql + ' RETURNING id', params)
        return cur.fetchone()['id']


def _log(acao, entidade, entidade_id=None, descricao=''):
    try:
        uid = current_user.id if current_user.is_authenticated else None
        unome = (current_user.nome or current_user.username) if current_user.is_authenticated else 'Sistema'
        _run(
            "INSERT INTO log_auditoria (usuario_id, usuario_nome, acao, entidade, entidade_id, descricao)"
            " VALUES (%s, %s, %s, %s, %s, %s)",
            (uid, unome, acao, entidade, entidade_id, descricao)
        )
        _run("DELETE FROM log_auditoria WHERE created_at < NOW() - INTERVAL '365 days'")
    except Exception:
        pass


def init_db():
    url = DATABASE_URL
    if url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
    conn.autocommit = True
    with conn.cursor() as cur:
        cur.execute("CREATE EXTENSION IF NOT EXISTS unaccent;")
        cur.execute('''
        CREATE TABLE IF NOT EXISTS usuario (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password_hash TEXT NOT NULL,
            nome TEXT,
            role TEXT DEFAULT 'operador'
        );
        CREATE TABLE IF NOT EXISTS estagiario (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            cpf TEXT UNIQUE NOT NULL,
            rg TEXT,
            data_nascimento TEXT,
            telefone TEXT,
            email TEXT,
            endereco TEXT,
            banco TEXT,
            agencia TEXT,
            conta TEXT,
            obs TEXT
        );
        CREATE TABLE IF NOT EXISTS empresa (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            cnpj TEXT,
            endereco TEXT,
            cidade TEXT DEFAULT \'Vitória da Conquista\',
            telefone TEXT,
            email TEXT,
            ramo TEXT,
            representante TEXT,
            cargo_representante TEXT,
            supervisor_nome TEXT,
            supervisor_cargo TEXT,
            supervisor_registro TEXT
        );
        CREATE TABLE IF NOT EXISTS ie (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            sigla TEXT,
            endereco TEXT,
            cidade TEXT DEFAULT \'Vitória da Conquista\',
            telefone TEXT,
            email TEXT,
            coordenador TEXT,
            coordenador_cargo TEXT
        );
        CREATE TABLE IF NOT EXISTS contrato (
            id SERIAL PRIMARY KEY,
            estagiario_id INTEGER NOT NULL REFERENCES estagiario(id),
            empresa_id INTEGER NOT NULL REFERENCES empresa(id),
            ie_id INTEGER NOT NULL REFERENCES ie(id),
            orientador TEXT DEFAULT \'Salmo Lima Costa\',
            supervisor_nome TEXT,
            supervisor_cargo TEXT,
            supervisor_registro TEXT,
            curso TEXT NOT NULL,
            tipo_estagio TEXT DEFAULT \'Não Obrigatório\',
            area_atuacao TEXT,
            ch_diaria INTEGER DEFAULT 6,
            ch_semanal INTEGER DEFAULT 30,
            data_inicio TEXT NOT NULL,
            data_fim TEXT NOT NULL,
            numero_contrato TEXT,
            bolsa REAL,
            taxa REAL,
            aux_transporte REAL,
            atividades TEXT,
            obs TEXT,
            num_relatorio INTEGER DEFAULT 1,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        ''')
        cur.execute("ALTER TABLE contrato ADD COLUMN IF NOT EXISTS jornada TEXT;")
        cur.execute("ALTER TABLE contrato ADD COLUMN IF NOT EXISTS data_encerramento TEXT;")
        cur.execute("ALTER TABLE estagiario ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'ativo';")
        cur.execute("UPDATE estagiario SET status = 'ativo' WHERE status IS NULL;")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS status TEXT DEFAULT 'ativo';")
        cur.execute("UPDATE empresa SET status = 'ativo' WHERE status IS NULL;")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS cpf_representante TEXT;")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS nome_fantasia TEXT;")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS bolsa_padrao REAL;")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS aux_transporte_padrao REAL;")
        cur.execute("ALTER TABLE contrato ADD COLUMN IF NOT EXISTS bolsa_tipo TEXT DEFAULT 'mensal';")
        cur.execute("ALTER TABLE estagiario ADD COLUMN IF NOT EXISTS semestre INTEGER;")
        cur.execute("ALTER TABLE estagiario ADD COLUMN IF NOT EXISTS tipo_ensino TEXT DEFAULT 'superior';")
        cur.execute("ALTER TABLE estagiario ADD COLUMN IF NOT EXISTS matricula TEXT;")
        cur.execute("ALTER TABLE estagiario ADD COLUMN IF NOT EXISTS ie_id INTEGER REFERENCES ie(id) ON DELETE SET NULL;")
        cur.execute("ALTER TABLE estagiario ADD COLUMN IF NOT EXISTS cidade TEXT;")
        cur.execute("ALTER TABLE estagiario ADD COLUMN IF NOT EXISTS estado TEXT;")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS estado TEXT;")
        cur.execute("ALTER TABLE ie ADD COLUMN IF NOT EXISTS estado TEXT;")
        cur.execute("ALTER TABLE ie ADD COLUMN IF NOT EXISTS cnpj TEXT;")
        cur.execute("ALTER TABLE ie ADD COLUMN IF NOT EXISTS representante_legal TEXT;")
        cur.execute("ALTER TABLE ie ADD COLUMN IF NOT EXISTS cargo_representante_legal TEXT;")
        cur.execute("ALTER TABLE ie ADD COLUMN IF NOT EXISTS signatario_tce TEXT DEFAULT 'coordenador';")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS ie_professor (
            id SERIAL PRIMARY KEY,
            ie_id INTEGER NOT NULL REFERENCES ie(id) ON DELETE CASCADE,
            nome TEXT NOT NULL,
            cargo TEXT,
            ordem INTEGER DEFAULT 0
        )""")
        cur.execute("ALTER TABLE contrato ADD COLUMN IF NOT EXISTS ie_professor_id INTEGER;")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS empresa_supervisor (
            id SERIAL PRIMARY KEY,
            empresa_id INTEGER NOT NULL REFERENCES empresa(id) ON DELETE CASCADE,
            nome TEXT NOT NULL,
            cargo TEXT,
            registro TEXT,
            ordem INTEGER DEFAULT 0
        )""")
        # Migra supervisores antigos (campo único) para a nova tabela
        cur.execute("""
            INSERT INTO empresa_supervisor (empresa_id, nome, cargo, registro, ordem)
            SELECT id, supervisor_nome, supervisor_cargo, supervisor_registro, 0
            FROM empresa
            WHERE supervisor_nome IS NOT NULL AND supervisor_nome <> ''
              AND id NOT IN (SELECT empresa_id FROM empresa_supervisor)
        """)
        cur.execute("""
        CREATE TABLE IF NOT EXISTS area_estagio (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            status TEXT DEFAULT 'ativo'
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS area_atividade (
            id SERIAL PRIMARY KEY,
            area_id INTEGER NOT NULL REFERENCES area_estagio(id) ON DELETE CASCADE,
            descricao TEXT NOT NULL,
            ordem INTEGER DEFAULT 0
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS aditivo (
            id SERIAL PRIMARY KEY,
            contrato_id INTEGER NOT NULL REFERENCES contrato(id) ON DELETE CASCADE,
            nova_data_fim TEXT,
            clausulas TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS relatorio_periodo (
            id SERIAL PRIMARY KEY,
            contrato_id INTEGER NOT NULL REFERENCES contrato(id) ON DELETE CASCADE,
            numero INTEGER NOT NULL,
            data_inicio TEXT NOT NULL,
            data_fim TEXT NOT NULL,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("ALTER TABLE usuario ADD COLUMN IF NOT EXISTS acesso_crm BOOLEAN DEFAULT FALSE;")
        cur.execute("ALTER TABLE usuario ADD COLUMN IF NOT EXISTS crm_role TEXT DEFAULT 'vendedor';")
        cur.execute("UPDATE usuario SET acesso_crm = TRUE WHERE role = 'admin';")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS crm_lead (
            id SERIAL PRIMARY KEY,
            empresa_nome TEXT NOT NULL,
            empresa_cnpj TEXT,
            cidade TEXT,
            segmento TEXT,
            vagas_estimadas INTEGER,
            etapa TEXT NOT NULL DEFAULT 'Lead Captado',
            origem TEXT,
            responsavel_id INTEGER REFERENCES usuario(id),
            contato_nome TEXT,
            contato_email TEXT,
            contato_whatsapp TEXT,
            obs TEXT,
            empresa_id INTEGER REFERENCES empresa(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS crm_interacao (
            id SERIAL PRIMARY KEY,
            lead_id INTEGER NOT NULL REFERENCES crm_lead(id) ON DELETE CASCADE,
            tipo TEXT NOT NULL,
            descricao TEXT,
            usuario_id INTEGER REFERENCES usuario(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS crm_implantacao (
            id SERIAL PRIMARY KEY,
            lead_id INTEGER NOT NULL REFERENCES crm_lead(id) ON DELETE CASCADE,
            empresa_id INTEGER REFERENCES empresa(id),
            responsavel_id INTEGER REFERENCES usuario(id),
            status TEXT NOT NULL DEFAULT 'em_andamento',
            obs TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS crm_implantacao_item (
            id SERIAL PRIMARY KEY,
            implantacao_id INTEGER NOT NULL REFERENCES crm_implantacao(id) ON DELETE CASCADE,
            titulo TEXT NOT NULL,
            ordem INTEGER DEFAULT 0,
            concluido BOOLEAN DEFAULT FALSE,
            concluido_em TIMESTAMP,
            concluido_por INTEGER REFERENCES usuario(id)
        )""")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS nps INTEGER;")
        cur.execute("ALTER TABLE empresa ADD COLUMN IF NOT EXISTS bairro TEXT;")
        cur.execute("ALTER TABLE candidato ADD COLUMN IF NOT EXISTS endereco TEXT;")
        cur.execute("ALTER TABLE candidato ADD COLUMN IF NOT EXISTS bairro TEXT;")
        cur.execute("ALTER TABLE ie ADD COLUMN IF NOT EXISTS data_vencimento_convenio DATE;")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS relacionamento_contato (
            id SERIAL PRIMARY KEY,
            entidade_tipo TEXT NOT NULL,
            entidade_id INTEGER NOT NULL,
            tipo TEXT NOT NULL,
            descricao TEXT,
            usuario_id INTEGER REFERENCES usuario(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS candidato (
            id SERIAL PRIMARY KEY,
            nome TEXT NOT NULL,
            cpf TEXT,
            email TEXT,
            whatsapp TEXT,
            data_nascimento DATE,
            cidade TEXT,
            estado TEXT,
            curso TEXT,
            semestre TEXT,
            ie_id INTEGER REFERENCES ie(id),
            disponibilidade TEXT,
            obs TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS vaga (
            id SERIAL PRIMARY KEY,
            empresa_id INTEGER REFERENCES empresa(id),
            area_id INTEGER REFERENCES area_estagio(id),
            titulo TEXT NOT NULL,
            descricao TEXT,
            requisitos TEXT,
            curso_desejado TEXT,
            nivel TEXT DEFAULT 'superior',
            carga_horaria INTEGER,
            bolsa NUMERIC(10,2),
            beneficios TEXT,
            status TEXT DEFAULT 'aberta',
            vagas_total INTEGER DEFAULT 1,
            data_limite DATE,
            responsavel_id INTEGER REFERENCES usuario(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS candidatura (
            id SERIAL PRIMARY KEY,
            vaga_id INTEGER NOT NULL REFERENCES vaga(id) ON DELETE CASCADE,
            candidato_id INTEGER NOT NULL REFERENCES candidato(id) ON DELETE CASCADE,
            status TEXT DEFAULT 'inscrito',
            obs TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            UNIQUE(vaga_id, candidato_id)
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS candidato_experiencia (
            id SERIAL PRIMARY KEY,
            candidato_id INTEGER NOT NULL REFERENCES candidato(id) ON DELETE CASCADE,
            empresa TEXT NOT NULL,
            periodo TEXT,
            funcao TEXT,
            ordem INTEGER DEFAULT 0
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS config (
            chave TEXT PRIMARY KEY,
            valor TEXT
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS log_auditoria (
            id SERIAL PRIMARY KEY,
            created_at TIMESTAMP DEFAULT NOW(),
            usuario_id INTEGER,
            usuario_nome TEXT,
            acao TEXT NOT NULL,
            entidade TEXT NOT NULL,
            entidade_id INTEGER,
            descricao TEXT
        )""")
        cur.execute("""
        CREATE TABLE IF NOT EXISTS prospecto (
            id SERIAL PRIMARY KEY,
            empresa_nome TEXT NOT NULL,
            cnpj TEXT,
            segmento TEXT,
            cnae_codigo TEXT,
            cnae_descricao TEXT,
            porte TEXT,
            cidade TEXT DEFAULT 'Vitória da Conquista',
            bairro TEXT,
            endereco TEXT,
            telefone TEXT,
            email TEXT,
            contato_nome TEXT,
            contato_cargo TEXT,
            site TEXT,
            vagas_estimadas INTEGER,
            obs TEXT,
            status TEXT DEFAULT 'novo',
            lead_id INTEGER REFERENCES crm_lead(id),
            responsavel_id INTEGER REFERENCES usuario(id),
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )""")
        for chave in ['seg_seguradora', 'seg_apolice', 'seg_coberturas', 'seg_vigencia',
                      'int_nome', 'int_cnpj', 'int_endereco', 'int_cidade', 'int_estado',
                      'int_representante', 'int_cargo']:
            cur.execute(
                "INSERT INTO config (chave, valor) VALUES (%s, '') ON CONFLICT (chave) DO NOTHING",
                (chave,))
        cur.execute("SELECT id FROM usuario WHERE username = 'salmo'")
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO usuario (username, password_hash, nome, role) VALUES (%s, %s, %s, %s)",
                ('salmo', generate_password_hash('ciclorh2026'), 'Salmo Lima Costa', 'admin')
            )
    conn.close()


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def _add_months(dt, months):
    month = dt.month - 1 + months
    year = dt.year + month // 12
    month = month % 12 + 1
    max_day = calendar.monthrange(year, month)[1]
    return dt.replace(year=year, month=month, day=min(dt.day, max_day))


def _split_periodos_6meses(inicio, fim):
    periodos = []
    atual = inicio
    while atual <= fim:
        chunk_fim = _add_months(atual, 6) - timedelta(days=1)
        if chunk_fim > fim:
            chunk_fim = fim
        periodos.append((atual, chunk_fim))
        atual = chunk_fim + timedelta(days=1)
    return periodos


def fmt_date(d):
    if not d:
        return ''
    if isinstance(d, str):
        try:
            d = datetime.strptime(d[:10], '%Y-%m-%d').date()
        except Exception:
            return d
    return d.strftime('%d/%m/%Y')


def calcular_status(data_fim_str):
    if not data_fim_str:
        return 'SEM DATA'
    try:
        fim = datetime.strptime(str(data_fim_str)[:10], '%Y-%m-%d').date()
        diff = (fim - date.today()).days
        if diff < 0:
            return 'VENCIDO'
        if diff <= 30:
            return f'VENCE EM {diff} DIAS'
        return 'ATIVO'
    except Exception:
        return 'ATIVO'


def br_currency(valor):
    try:
        v = float(valor or 0)
    except (TypeError, ValueError):
        return '0,00'
    s = '{:,.2f}'.format(v)
    return s.replace(',', 'X').replace('.', ',').replace('X', '.')


def num_extenso(valor):
    try:
        valor = round(float(valor), 2)
    except (TypeError, ValueError):
        return ''
    inteiro = int(valor)
    centavos = round((valor - inteiro) * 100)
    _un = ['', 'um', 'dois', 'três', 'quatro', 'cinco', 'seis', 'sete', 'oito', 'nove',
           'dez', 'onze', 'doze', 'treze', 'quatorze', 'quinze', 'dezesseis', 'dezessete', 'dezoito', 'dezenove']
    _dez = ['', '', 'vinte', 'trinta', 'quarenta', 'cinquenta', 'sessenta', 'setenta', 'oitenta', 'noventa']
    _cent = ['', 'cem', 'duzentos', 'trezentos', 'quatrocentos', 'quinhentos',
             'seiscentos', 'setecentos', 'oitocentos', 'novecentos']

    def dez(n):
        if n < 20:
            return _un[n]
        d, u = _dez[n // 10], _un[n % 10]
        return d + (' e ' + u if u else '')

    def cent(n):
        c, r = n // 100, n % 100
        if c == 1 and r == 0:
            return 'cem'
        base = 'cento' if c == 1 else _cent[c]
        return base + (' e ' + dez(r) if r else '')

    def grupo(n):
        return cent(n) if n >= 100 else dez(n)

    def por_extenso(n):
        if n == 0:
            return 'zero'
        partes = []
        if n >= 1000000:
            m = n // 1000000
            partes.append(grupo(m) + (' milhão' if m == 1 else ' milhões'))
            n %= 1000000
        if n >= 1000:
            m = n // 1000
            partes.append(('um' if m == 1 else grupo(m)) + ' mil')
            n %= 1000
        if n > 0:
            partes.append(grupo(n))
        return ' e '.join(partes)

    partes = []
    if inteiro > 0:
        partes.append(por_extenso(inteiro) + (' real' if inteiro == 1 else ' reais'))
    if centavos > 0:
        partes.append(dez(centavos) + (' centavo' if centavos == 1 else ' centavos'))
    return ' e '.join(partes) if partes else 'zero reais'


def formatar_jornada(jornada_json):
    if not jornada_json:
        return ''
    try:
        j = json.loads(jornada_json)
    except Exception:
        return ''

    DIAS_KEYS = ['dom', 'seg', 'ter', 'qua', 'qui', 'sex', 'sab']
    DIAS_ACUS = {
        'dom': 'aos domingos', 'seg': 'às segundas-feiras', 'ter': 'às terças-feiras',
        'qua': 'às quartas-feiras', 'qui': 'às quintas-feiras', 'sex': 'às sextas-feiras',
        'sab': 'aos sábados',
    }
    DIAS_DE = {
        'dom': 'domingo', 'seg': 'segunda', 'ter': 'terça', 'qua': 'quarta',
        'qui': 'quinta', 'sex': 'sexta', 'sab': 'sábado',
    }
    DIAS_ATE = {
        'dom': 'domingo', 'seg': 'segunda-feira', 'ter': 'terça-feira', 'qua': 'quarta-feira',
        'qui': 'quinta-feira', 'sex': 'sexta-feira', 'sab': 'sábado',
    }

    def dia_horarios(d):
        h = []
        for periodo in ['mat', 'ves', 'not']:
            ini = d.get(f'{periodo}_ini', '').strip()
            fim = d.get(f'{periodo}_fim', '').strip()
            if ini and fim:
                h.append((ini, fim))
        return tuple(h)

    schedule_groups = {}
    for k in DIAS_KEYS:
        if k not in j:
            continue
        h = dia_horarios(j[k])
        if not h:
            continue
        if h not in schedule_groups:
            schedule_groups[h] = []
        schedule_groups[h].append(k)

    if not schedule_groups:
        return ''

    parts = []
    for horarios, dias in schedule_groups.items():
        dias_sorted = sorted(dias, key=lambda d: DIAS_KEYS.index(d))
        n = len(dias_sorted)
        indices = [DIAS_KEYS.index(d) for d in dias_sorted]
        is_consecutive = (n > 1 and indices[-1] - indices[0] == n - 1)

        if is_consecutive and n >= 2:
            dia_fmt = f'de {DIAS_DE[dias_sorted[0]]} a {DIAS_ATE[dias_sorted[-1]]}'
        elif n == 1:
            dia_fmt = DIAS_ACUS[dias_sorted[0]]
        elif n == 2:
            dia_fmt = f'{DIAS_ACUS[dias_sorted[0]]} e {DIAS_ACUS[dias_sorted[1]]}'
        else:
            dia_fmt = ', '.join(DIAS_ACUS[d] for d in dias_sorted[:-1]) + f' e {DIAS_ACUS[dias_sorted[-1]]}'

        horario_fmt = ' e '.join(f'das {ini} às {fim}' for ini, fim in horarios)
        parts.append(f'{dia_fmt} {horario_fmt}')

    if not parts:
        return ''
    result = '; '.join(parts)
    return result[0].upper() + result[1:] + '.'


def calcular_ch_jornada(jornada_json):
    if not jornada_json:
        return None, None
    try:
        j = json.loads(jornada_json)
    except Exception:
        return None, None
    dias_keys = ['dom', 'seg', 'ter', 'qua', 'qui', 'sex', 'sab']
    total_semanal = 0.0
    dias_horas = []
    for dia in dias_keys:
        if dia not in j:
            continue
        dd = j[dia]
        horas_dia = 0.0
        for periodo in ['mat', 'ves', 'not']:
            ini = dd.get(f'{periodo}_ini', '').strip()
            fim = dd.get(f'{periodo}_fim', '').strip()
            if ini and fim:
                try:
                    hi = int(ini[:2]) + int(ini[3:5]) / 60
                    hf = int(fim[:2]) + int(fim[3:5]) / 60
                    if hf > hi:
                        horas_dia += hf - hi
                except Exception:
                    pass
        if horas_dia > 0:
            dias_horas.append(horas_dia)
            total_semanal += horas_dia
    if not dias_horas:
        return None, None
    return round(max(dias_horas), 1), round(total_semanal, 1)


def _build_jornada_json():
    dias = ['dom', 'seg', 'ter', 'qua', 'qui', 'sex', 'sab']
    jornada = {}
    for dia in dias:
        dd = {}
        for periodo in ['mat', 'ves', 'not']:
            ini = request.form.get(f'{dia}_{periodo}_ini', '').strip()
            fim = request.form.get(f'{dia}_{periodo}_fim', '').strip()
            if ini or fim:
                dd[f'{periodo}_ini'] = ini
                dd[f'{periodo}_fim'] = fim
        if dd:
            jornada[dia] = dd
    return json.dumps(jornada, ensure_ascii=False) if jornada else None


def _get_config():
    try:
        rows = _q("SELECT chave, valor FROM config")
        return {r['chave']: r['valor'] or '' for r in rows}
    except Exception:
        return {}


def _fmt_semestre(semestre, tipo_ensino):
    if not semestre:
        return ''
    if tipo_ensino == 'medio':
        return f'{semestre}º Ano'
    return f'{semestre}º Semestre'


_semestre_atualizado = set()


def _atualizar_semestres_auto():
    hoje = date.today()
    if hoje.month not in (1, 7):
        return
    key = (hoje.year, hoje.month)
    if key in _semestre_atualizado:
        return
    _semestre_atualizado.add(key)
    chave_cfg = f'sem_upd_{hoje.year}_{hoje.month:02d}'
    try:
        cfg = _get_config()
        if cfg.get(chave_cfg):
            return
        if hoje.month == 7:
            _run("""UPDATE estagiario SET semestre = LEAST(semestre + 1, 10)
                    WHERE tipo_ensino IN ('superior', 'tecnico')
                    AND semestre IS NOT NULL AND semestre < 10""")
        else:
            _run("""UPDATE estagiario SET semestre = CASE
                        WHEN tipo_ensino = 'medio' THEN LEAST(semestre + 1, 3)
                        ELSE LEAST(semestre + 1, 10)
                    END WHERE semestre IS NOT NULL""")
        _run("INSERT INTO config (chave, valor) VALUES (%s, %s) ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
             (chave_cfg, hoje.isoformat()))
    except Exception:
        pass


@app.before_request
def _antes_da_requisicao():
    _atualizar_semestres_auto()


app.jinja_env.globals.update(fmt_date=fmt_date, calcular_status=calcular_status,
                             fmt_semestre=_fmt_semestre)


@app.context_processor
def inject_today():
    return {'today': date.today()}
app.jinja_env.filters['from_json'] = lambda s: json.loads(s) if s else {}


@app.context_processor
def inject_pending():
    pending = 0
    if current_user.is_authenticated and current_user.is_admin:
        try:
            r1 = _q("SELECT COUNT(*) n FROM estagiario WHERE status='pendente'", one=True)
            r2 = _q("SELECT COUNT(*) n FROM empresa WHERE status='pendente'", one=True)
            pending = (r1['n'] if r1 else 0) + (r2['n'] if r2 else 0)
        except Exception:
            pending = 0
    return {'pending_count': pending}


# ─── AUTH ─────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    if request.method == 'POST':
        username = request.form.get('username', '').strip().lower()
        senha = request.form.get('senha', '')
        row = _q("SELECT * FROM usuario WHERE username = %s", (username,), one=True)
        if row and check_password_hash(row['password_hash'], senha):
            user = User(row['id'], row['username'], row['nome'], row['role'],
                        row.get('acesso_crm', False), row.get('crm_role', 'vendedor'))
            login_user(user, remember=True)
            _log('login', 'sistema', None, f'Login: {username}')
            return redirect(request.args.get('next') or url_for('index'))
        flash('Usuário ou senha incorretos.', 'danger')
    return render_template('auth/login.html')


@app.route('/logout')
@login_required
def logout():
    _log('logout', 'sistema', None, f'Logout: {current_user.username}')
    logout_user()
    return redirect(url_for('login'))


# ─── ADMIN — USUÁRIOS ─────────────────────────────────────────────────────────

@app.route('/admin/usuarios')
@admin_required
def admin_usuarios():
    rows = _q("SELECT id, username, nome, role, acesso_crm, crm_role FROM usuario ORDER BY role DESC, nome")
    return render_template('admin/usuarios.html', usuarios=rows)


@app.route('/admin/usuarios/novo', methods=['GET', 'POST'])
@admin_required
def admin_usuario_novo():
    if request.method == 'POST':
        username = request.form['username'].strip().lower()
        senha = request.form['senha']
        nome = request.form.get('nome', '').strip()
        role = request.form.get('role', 'operador')
        if len(senha) < 6:
            flash('Senha deve ter no mínimo 6 caracteres.', 'danger')
        else:
            try:
                acesso_crm_n = 'acesso_crm' in request.form or role == 'admin'
                crm_role_n = request.form.get('crm_role', 'vendedor')
                _ins("INSERT INTO usuario (username, password_hash, nome, role, acesso_crm, crm_role) VALUES (%s,%s,%s,%s,%s,%s)",
                     (username, generate_password_hash(senha), nome, role, acesso_crm_n, crm_role_n))
                _log('criar', 'usuario', None, f'Criou usuário: {username} ({role})')
                flash(f'Usuário "{username}" criado!', 'success')
                return redirect(url_for('admin_usuarios'))
            except psycopg2.errors.UniqueViolation:
                flash('Nome de usuário já existe.', 'danger')
    return render_template('admin/usuario_form.html', u=None)


@app.route('/admin/usuarios/<int:id>/editar', methods=['GET', 'POST'])
@admin_required
def admin_usuario_editar(id):
    u = _q("SELECT * FROM usuario WHERE id = %s", (id,), one=True)
    if not u:
        abort(404)
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        role = request.form.get('role', 'operador')
        acesso_crm = 'acesso_crm' in request.form or role == 'admin'
        crm_role = request.form.get('crm_role', 'vendedor')
        _run("UPDATE usuario SET nome=%s, role=%s, acesso_crm=%s, crm_role=%s WHERE id=%s",
             (nome, role, acesso_crm, crm_role, id))
        senha = request.form.get('senha', '').strip()
        if senha:
            if len(senha) < 6:
                flash('Senha deve ter no mínimo 6 caracteres.', 'danger')
                return render_template('admin/usuario_form.html', u=u)
            _run("UPDATE usuario SET password_hash = %s WHERE id = %s",
                 (generate_password_hash(senha), id))
        _log('editar', 'usuario', id, f'Editou usuário ID {id} ({u["username"]})')
        flash('Usuário atualizado!', 'success')
        return redirect(url_for('admin_usuarios'))
    return render_template('admin/usuario_form.html', u=u)


@app.route('/admin/usuarios/<int:id>/excluir')
@admin_required
def admin_usuario_excluir(id):
    if str(id) == current_user.id:
        flash('Você não pode excluir sua própria conta.', 'danger')
        return redirect(url_for('admin_usuarios'))
    reg = _q("SELECT username FROM usuario WHERE id = %s", (id,), one=True)
    _run("DELETE FROM usuario WHERE id = %s", (id,))
    _log('excluir', 'usuario', id, f'Excluiu usuário: {reg["username"] if reg else id}')
    flash('Usuário excluído.', 'warning')
    return redirect(url_for('admin_usuarios'))


# ─── ADMIN — IMPORTAR PLANILHA ────────────────────────────────────────────────

@app.route('/admin/importar', methods=['GET', 'POST'])
@admin_required
def admin_importar():
    resultado = None
    if request.method == 'POST':
        arquivo = request.files.get('arquivo')
        if not arquivo or not arquivo.filename.endswith(('.xlsx', '.xlsm')):
            flash('Envie um arquivo .xlsx ou .xlsm válido.', 'danger')
            return render_template('admin/importar.html', resultado=None)
        import tempfile, importar as imp_mod
        with tempfile.NamedTemporaryFile(suffix='.xlsm', delete=False) as tmp:
            arquivo.save(tmp.name)
            resultado = imp_mod.run_from_file(tmp.name, DATABASE_URL)
        os.unlink(tmp.name)
        flash(f'Importação concluída: {resultado["ok"]} contratos importados.', 'success')
    return render_template('admin/importar.html', resultado=resultado)


# ─── LANDING PAGE PÚBLICA ─────────────────────────────────────────────────────

@app.route('/landing')
def landing():
    return render_template('landing.html')


# ─── DASHBOARD ────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def index():
    hoje = date.today().isoformat()
    d30 = (date.today() + timedelta(days=30)).isoformat()

    # CTE que calcula a data efetiva de fim (considera aditivos de prorrogação)
    cte = """
        WITH ef AS (
            SELECT c.id,
                   COALESCE(
                       c.data_encerramento,
                       (SELECT a.nova_data_fim FROM aditivo a
                        WHERE a.contrato_id = c.id AND a.nova_data_fim IS NOT NULL AND a.nova_data_fim <> ''
                        ORDER BY a.created_at DESC LIMIT 1),
                       c.data_fim
                   ) AS data_efetiva
            FROM contrato c
        )
    """

    # Contratos ativos (data efetiva >= hoje e sem encerramento)
    total = _q(cte + """
        SELECT COUNT(*) AS n FROM contrato c
        JOIN ef ON ef.id = c.id
        WHERE c.data_encerramento IS NULL AND ef.data_efetiva >= %s
    """, (hoje,), one=True)['n']

    # Vencendo em 30 dias (ainda ativos)
    vencendo = _q(cte + """
        SELECT c.*, e.nome est_nome, emp.nome emp_nome,
               COALESCE(emp.nome_fantasia, emp.nome) emp_display, ef.data_efetiva
        FROM contrato c
        JOIN estagiario e ON e.id = c.estagiario_id
        JOIN empresa emp ON emp.id = c.empresa_id
        JOIN ef ON ef.id = c.id
        WHERE c.data_encerramento IS NULL
          AND ef.data_efetiva >= %s AND ef.data_efetiva <= %s
        ORDER BY ef.data_efetiva
    """, (hoje, d30))

    # Vencidos sem confirmação (precisam de ação)
    pendentes = _q(cte + """
        SELECT c.*, e.nome est_nome, emp.nome emp_nome,
               COALESCE(emp.nome_fantasia, emp.nome) emp_display, ef.data_efetiva
        FROM contrato c
        JOIN estagiario e ON e.id = c.estagiario_id
        JOIN empresa emp ON emp.id = c.empresa_id
        JOIN ef ON ef.id = c.id
        WHERE c.data_encerramento IS NULL AND ef.data_efetiva < %s
        ORDER BY ef.data_efetiva DESC
    """, (hoje,))

    # Contratos recentes ativos
    recentes = _q(cte + """
        SELECT c.*, e.nome est_nome, emp.nome emp_nome,
               COALESCE(emp.nome_fantasia, emp.nome) emp_display, ef.data_efetiva
        FROM contrato c
        JOIN estagiario e ON e.id = c.estagiario_id
        JOIN empresa emp ON emp.id = c.empresa_id
        JOIN ef ON ef.id = c.id
        WHERE c.data_encerramento IS NULL AND ef.data_efetiva >= %s
        ORDER BY c.created_at DESC LIMIT 10
    """, (hoje,))

    total_est = _q("SELECT COUNT(*) AS n FROM estagiario", one=True)['n']
    total_emp = _q("SELECT COUNT(*) AS n FROM empresa", one=True)['n']
    total_ie = _q("SELECT COUNT(*) AS n FROM ie", one=True)['n']
    return render_template('index.html', total=total, vencendo=vencendo, recentes=recentes,
                           pendentes=pendentes,
                           total_est=total_est, total_emp=total_emp, total_ie=total_ie)


@app.route('/dashboard')
@login_required
def dashboard():
    hoje = date.today()
    hoje_s = hoje.isoformat()
    d30_s = (hoje + timedelta(days=30)).isoformat()
    mes_ini = hoje.replace(day=1).isoformat()

    cte = """WITH ef AS (
        SELECT c.id,
               COALESCE(c.data_encerramento,
                   (SELECT a.nova_data_fim FROM aditivo a
                    WHERE a.contrato_id = c.id AND a.nova_data_fim IS NOT NULL AND a.nova_data_fim <> ''
                    ORDER BY a.created_at DESC LIMIT 1),
                   c.data_fim) AS data_efetiva
        FROM contrato c)"""

    # ── Estágios ──────────────────────────────────────────────────────────────
    ativos = _q(cte + " SELECT COUNT(*) n FROM contrato c JOIN ef ON ef.id=c.id"
                " WHERE c.data_encerramento IS NULL AND ef.data_efetiva >= %s",
                (hoje_s,), one=True)['n']
    vencendo_30 = _q(cte + " SELECT COUNT(*) n FROM contrato c JOIN ef ON ef.id=c.id"
                     " WHERE c.data_encerramento IS NULL AND ef.data_efetiva >= %s AND ef.data_efetiva <= %s",
                     (hoje_s, d30_s), one=True)['n']
    encerrados_mes = _q("SELECT COUNT(*) n FROM contrato WHERE data_encerramento >= %s", (mes_ini,), one=True)['n']
    novos_mes = _q("SELECT COUNT(*) n FROM contrato WHERE created_at >= %s", (mes_ini,), one=True)['n']
    vencidos_sem_tre = _q(cte + " SELECT COUNT(*) n FROM contrato c JOIN ef ON ef.id=c.id"
                          " WHERE c.data_encerramento IS NULL AND ef.data_efetiva < %s",
                          (hoje_s,), one=True)['n']

    # ── CRM ───────────────────────────────────────────────────────────────────
    crm_total = _q("SELECT COUNT(*) n FROM crm_lead", one=True)['n']
    crm_ativos = _q("SELECT COUNT(*) n FROM crm_lead WHERE etapa='Cliente Ativo'", one=True)['n']
    crm_taxa = round(crm_ativos / crm_total * 100, 1) if crm_total else 0
    crm_novos_mes = _q("SELECT COUNT(*) n FROM crm_lead WHERE created_at >= %s", (mes_ini,), one=True)['n']
    crm_por_etapa = _q("""SELECT etapa, COUNT(*) as n FROM crm_lead GROUP BY etapa""")
    crm_etapa_map = {r['etapa']: r['n'] for r in crm_por_etapa}

    # ── Implantações ──────────────────────────────────────────────────────────
    imp_andamento = _q("SELECT COUNT(*) n FROM crm_implantacao WHERE status='em_andamento'", one=True)['n']
    imp_concluidas = _q("SELECT COUNT(*) n FROM crm_implantacao WHERE status='concluida'", one=True)['n']
    imp_progresso = _q("""SELECT AVG(pct) as media FROM (
        SELECT i.id, CASE WHEN COUNT(it.id)=0 THEN 0
               ELSE ROUND(COUNT(CASE WHEN it.concluido THEN 1 END)::numeric / COUNT(it.id) * 100)
               END as pct
        FROM crm_implantacao i
        LEFT JOIN crm_implantacao_item it ON it.implantacao_id=i.id
        WHERE i.status='em_andamento' GROUP BY i.id) sub""", one=True)['media']
    imp_progresso = round(float(imp_progresso)) if imp_progresso else 0
    imp_paradas = _q("""SELECT COUNT(*) n FROM crm_implantacao
                        WHERE status='em_andamento' AND updated_at < NOW() - INTERVAL '30 days'""",
                     one=True)['n']

    # ── Vagas ─────────────────────────────────────────────────────────────────
    vagas_abertas = _q("SELECT COUNT(*) n FROM vaga WHERE status IN ('aberta','em_selecao')", one=True)['n']
    vagas_candidatos = _q("SELECT COUNT(*) n FROM candidatura WHERE status='aprovado'", one=True)['n']
    total_candidatos = _q("SELECT COUNT(*) n FROM candidato", one=True)['n']

    # ── Alertas ───────────────────────────────────────────────────────────────
    alertas = []
    if vencidos_sem_tre:
        alertas.append({'tipo': 'danger', 'icone': 'bi-exclamation-triangle',
                        'msg': f'{vencidos_sem_tre} contrato(s) vencido(s) sem TRE emitido.',
                        'link': url_for('contratos'), 'link_label': 'Ver contratos'})
    if imp_paradas:
        alertas.append({'tipo': 'warning', 'icone': 'bi-hourglass-split',
                        'msg': f'{imp_paradas} implantação(ões) sem atualização há mais de 30 dias.',
                        'link': url_for('crm_implantacao_lista'), 'link_label': 'Ver implantações'})
    if vagas_candidatos:
        alertas.append({'tipo': 'info', 'icone': 'bi-person-check',
                        'msg': f'{vagas_candidatos} candidato(s) aprovado(s) aguardando contrato.',
                        'link': url_for('vagas_lista'), 'link_label': 'Ver vagas'})
    if vencendo_30:
        alertas.append({'tipo': 'warning', 'icone': 'bi-calendar-x',
                        'msg': f'{vencendo_30} contrato(s) vencem nos próximos 30 dias.',
                        'link': url_for('relatorio_vencimentos'), 'link_label': 'Ver TCEs'})

    ies_vencendo = _q("""SELECT id, nome, sigla, data_vencimento_convenio FROM ie
                         WHERE data_vencimento_convenio IS NOT NULL
                           AND data_vencimento_convenio <= CURRENT_DATE + INTERVAL '60 days'
                           AND data_vencimento_convenio >= CURRENT_DATE
                         ORDER BY data_vencimento_convenio""")
    for ie in ies_vencendo:
        dias = (ie['data_vencimento_convenio'] - date.today()).days
        alertas.append({'tipo': 'warning', 'icone': 'bi-mortarboard',
                        'msg': f'Convênio com {ie["sigla"] or ie["nome"]} vence em {dias} dia(s).',
                        'link': url_for('ie_detalhe', id=ie['id']), 'link_label': 'Ver IE'})

    return render_template('dashboard.html',
                           ativos=ativos, vencendo_30=vencendo_30,
                           encerrados_mes=encerrados_mes, novos_mes=novos_mes,
                           vencidos_sem_tre=vencidos_sem_tre,
                           crm_total=crm_total, crm_ativos=crm_ativos,
                           crm_taxa=crm_taxa, crm_novos_mes=crm_novos_mes,
                           crm_etapa_map=crm_etapa_map, etapas_crm=ETAPAS_CRM,
                           cores_crm=ETAPAS_CRM_COR,
                           imp_andamento=imp_andamento, imp_concluidas=imp_concluidas,
                           imp_progresso=imp_progresso, imp_paradas=imp_paradas,
                           vagas_abertas=vagas_abertas, vagas_candidatos=vagas_candidatos,
                           total_candidatos=total_candidatos,
                           alertas=alertas)


@app.route('/contratos/<int:id>/encerrar', methods=['POST'])
@login_required
def contrato_encerrar(id):
    c = _q("SELECT * FROM contrato WHERE id = %s", (id,), one=True)
    if not c:
        abort(404)
    # Usa a data efetiva (último aditivo ou data_fim original)
    data_efetiva = _q("""
        SELECT COALESCE(
            (SELECT nova_data_fim FROM aditivo
             WHERE contrato_id = %s AND nova_data_fim IS NOT NULL AND nova_data_fim <> ''
             ORDER BY created_at DESC LIMIT 1),
            %s
        ) AS dt
    """, (id, c['data_fim']), one=True)['dt']
    _run("UPDATE contrato SET data_encerramento = %s WHERE id = %s", (data_efetiva, id))
    est = _q("SELECT nome FROM estagiario WHERE id = %s", (c['estagiario_id'],), one=True)
    _log('encerrar', 'contrato', id, f'Encerrou contrato: {est["nome"] if est else id}')
    flash(f'Contrato encerrado em {fmt_date(data_efetiva)}.', 'success')
    return redirect(url_for('index'))


# ─── ESTAGIÁRIOS ──────────────────────────────────────────────────────────────

@app.route('/estagiarios')
@login_required
def estagiarios():
    q = request.args.get('q', '')
    if q:
        rows = _q("""SELECT e.*,
                     (SELECT COUNT(*) FROM contrato WHERE estagiario_id = e.id) qtd_contratos
                     FROM estagiario e
                     WHERE (e.nome ILIKE %s OR e.cpf ILIKE %s) AND e.status = 'ativo' ORDER BY e.nome""",
                  (f'%{q}%', f'%{q}%'))
    else:
        rows = _q("""SELECT e.*,
                     (SELECT COUNT(*) FROM contrato WHERE estagiario_id = e.id) qtd_contratos
                     FROM estagiario e WHERE e.status = 'ativo' ORDER BY e.nome""")
    return render_template('estagiarios/lista.html', estagiarios=rows, q=q)


@app.route('/estagiarios/novo', methods=['GET', 'POST'])
@login_required
def estagiario_novo():
    ies = _q("SELECT id, nome, sigla FROM ie ORDER BY nome")
    if request.method == 'POST':
        try:
            _ins("""INSERT INTO estagiario
                    (nome,cpf,rg,data_nascimento,telefone,email,endereco,cidade,estado,banco,agencia,conta,obs,
                     tipo_ensino,semestre,matricula,ie_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                 (request.form['nome'], request.form['cpf'],
                  request.form.get('rg'), request.form.get('data_nascimento') or None,
                  request.form.get('telefone'), request.form.get('email'),
                  request.form.get('endereco'), request.form.get('cidade') or None,
                  request.form.get('estado') or None,
                  request.form.get('banco'),
                  request.form.get('agencia'), request.form.get('conta'),
                  request.form.get('obs'),
                  request.form.get('tipo_ensino', 'superior'),
                  request.form.get('semestre') or None,
                  request.form.get('matricula') or None,
                  request.form.get('ie_id') or None))
            _log('criar', 'estagiario', None, f'Criou estagiário: {request.form["nome"]} (CPF: {request.form["cpf"]})')
            flash('Estagiário cadastrado!', 'success')
            return redirect(url_for('estagiarios'))
        except psycopg2.errors.UniqueViolation:
            flash('CPF já cadastrado!', 'danger')
    return render_template('estagiarios/form.html', e=None, ies=ies)


@app.route('/estagiarios/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def estagiario_editar(id):
    e = _q("SELECT * FROM estagiario WHERE id = %s", (id,), one=True)
    if not e:
        abort(404)
    ies = _q("SELECT id, nome, sigla FROM ie ORDER BY nome")
    if request.method == 'POST':
        _run("""UPDATE estagiario SET
                nome=%s,cpf=%s,rg=%s,data_nascimento=%s,telefone=%s,email=%s,
                endereco=%s,cidade=%s,estado=%s,banco=%s,agencia=%s,conta=%s,obs=%s,
                tipo_ensino=%s,semestre=%s,matricula=%s,ie_id=%s WHERE id=%s""",
             (request.form['nome'], request.form['cpf'],
              request.form.get('rg'), request.form.get('data_nascimento') or None,
              request.form.get('telefone'), request.form.get('email'),
              request.form.get('endereco'), request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('banco'),
              request.form.get('agencia'), request.form.get('conta'),
              request.form.get('obs'),
              request.form.get('tipo_ensino', 'superior'),
              request.form.get('semestre') or None,
              request.form.get('matricula') or None,
              request.form.get('ie_id') or None,
              id))
        _log('editar', 'estagiario', id, f'Editou estagiário: {request.form["nome"]}')
        flash('Atualizado!', 'success')
        return redirect(url_for('estagiarios'))
    return render_template('estagiarios/form.html', e=e, ies=ies)


@app.route('/estagiarios/<int:id>/excluir')
@login_required
def estagiario_excluir(id):
    reg = _q("SELECT nome FROM estagiario WHERE id = %s", (id,), one=True)
    _run("DELETE FROM estagiario WHERE id = %s", (id,))
    _log('excluir', 'estagiario', id, f'Excluiu estagiário: {reg["nome"] if reg else id}')
    flash('Excluído.', 'warning')
    return redirect(url_for('estagiarios'))


# ─── EMPRESAS ─────────────────────────────────────────────────────────────────

@app.route('/empresas')
@login_required
def empresas():
    q = request.args.get('q', '')
    aba = request.args.get('aba', 'ativas')

    ativas_cond = "EXISTS (SELECT 1 FROM contrato WHERE empresa_id=emp.id AND data_encerramento IS NULL)"
    inativas_cond = "NOT EXISTS (SELECT 1 FROM contrato WHERE empresa_id=emp.id AND data_encerramento IS NULL)"

    base_sql = """SELECT emp.*,
                 (SELECT COUNT(*) FROM contrato WHERE empresa_id = emp.id AND data_encerramento IS NULL) qtd_contratos,
                 (SELECT COUNT(*) FROM empresa_supervisor WHERE empresa_id = emp.id) qtd_supervisores,
                 (SELECT nome FROM empresa_supervisor WHERE empresa_id = emp.id ORDER BY ordem, id LIMIT 1) primeiro_supervisor
                 FROM empresa emp WHERE emp.status = 'ativo'"""

    extra = f" AND {ativas_cond if aba == 'ativas' else inativas_cond}"

    if q:
        rows = _q(base_sql + extra + " AND (emp.nome ILIKE %s OR emp.cnpj ILIKE %s OR emp.nome_fantasia ILIKE %s) ORDER BY emp.nome",
                  (f'%{q}%', f'%{q}%', f'%{q}%'))
    else:
        rows = _q(base_sql + extra + " ORDER BY emp.nome")

    total_ativas = _q("SELECT COUNT(*) n FROM empresa emp WHERE emp.status='ativo' AND " + ativas_cond, one=True)['n']
    total_inativas = _q("SELECT COUNT(*) n FROM empresa emp WHERE emp.status='ativo' AND " + inativas_cond, one=True)['n']

    return render_template('empresas/lista.html', empresas=rows, q=q, aba=aba,
                           total_ativas=total_ativas, total_inativas=total_inativas)


@app.route('/empresas/nova', methods=['GET', 'POST'])
@login_required
def empresa_nova():
    if request.method == 'POST':
        emp_id = _ins("""INSERT INTO empresa
                (nome,nome_fantasia,cnpj,endereco,bairro,cidade,estado,telefone,email,ramo,
                 representante,cargo_representante,cpf_representante,
                 bolsa_padrao,aux_transporte_padrao)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (request.form['nome'], request.form.get('nome_fantasia') or None,
              request.form.get('cnpj'),
              request.form.get('endereco'), request.form.get('bairro') or None,
              request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('telefone'), request.form.get('email'),
              request.form.get('ramo'), request.form.get('representante'),
              request.form.get('cargo_representante'), request.form.get('cpf_representante'),
              request.form.get('bolsa_padrao') or None,
              request.form.get('aux_transporte_padrao') or None))
        for i, nome_sup in enumerate(request.form.getlist('sup_nome[]')):
            if nome_sup.strip():
                cargos = request.form.getlist('sup_cargo[]')
                regs = request.form.getlist('sup_registro[]')
                _run("INSERT INTO empresa_supervisor (empresa_id,nome,cargo,registro,ordem) VALUES (%s,%s,%s,%s,%s)",
                     (emp_id, nome_sup.strip(),
                      cargos[i].strip() if i < len(cargos) else None,
                      regs[i].strip() if i < len(regs) else None, i))
        _log('criar', 'empresa', emp_id, f'Criou empresa: {request.form["nome"]}')
        flash('Empresa cadastrada!', 'success')
        return redirect(url_for('empresas'))
    return render_template('empresas/form.html', emp=None, supervisores=[])


@app.route('/empresas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def empresa_editar(id):
    emp = _q("SELECT * FROM empresa WHERE id = %s", (id,), one=True)
    if not emp:
        abort(404)
    if request.method == 'POST':
        _run("""UPDATE empresa SET
                nome=%s,nome_fantasia=%s,cnpj=%s,endereco=%s,bairro=%s,cidade=%s,estado=%s,telefone=%s,email=%s,ramo=%s,
                representante=%s,cargo_representante=%s,cpf_representante=%s,
                bolsa_padrao=%s,aux_transporte_padrao=%s WHERE id=%s""",
             (request.form['nome'], request.form.get('nome_fantasia') or None,
              request.form.get('cnpj'),
              request.form.get('endereco'), request.form.get('bairro') or None,
              request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('telefone'), request.form.get('email'),
              request.form.get('ramo'), request.form.get('representante'),
              request.form.get('cargo_representante'), request.form.get('cpf_representante'),
              request.form.get('bolsa_padrao') or None,
              request.form.get('aux_transporte_padrao') or None, id))
        _run("DELETE FROM empresa_supervisor WHERE empresa_id = %s", (id,))
        for i, nome_sup in enumerate(request.form.getlist('sup_nome[]')):
            if nome_sup.strip():
                cargos = request.form.getlist('sup_cargo[]')
                regs = request.form.getlist('sup_registro[]')
                _run("INSERT INTO empresa_supervisor (empresa_id,nome,cargo,registro,ordem) VALUES (%s,%s,%s,%s,%s)",
                     (id, nome_sup.strip(),
                      cargos[i].strip() if i < len(cargos) else None,
                      regs[i].strip() if i < len(regs) else None, i))
        _log('editar', 'empresa', id, f'Editou empresa: {request.form["nome"]}')
        flash('Atualizada!', 'success')
        return redirect(url_for('empresas'))
    supervisores = _q("SELECT * FROM empresa_supervisor WHERE empresa_id = %s ORDER BY ordem, id", (id,))
    return render_template('empresas/form.html', emp=emp, supervisores=supervisores)


@app.route('/empresas/<int:id>/excluir')
@login_required
def empresa_excluir(id):
    reg = _q("SELECT nome FROM empresa WHERE id = %s", (id,), one=True)
    _run("DELETE FROM empresa WHERE id = %s", (id,))
    _log('excluir', 'empresa', id, f'Excluiu empresa: {reg["nome"] if reg else id}')
    flash('Excluída.', 'warning')
    return redirect(url_for('empresas'))


@app.route('/empresas/<int:id>')
@login_required
def empresa_detalhe(id):
    emp = _q("""SELECT emp.*,
                (SELECT COUNT(*) FROM contrato WHERE empresa_id=emp.id AND data_encerramento IS NULL) contratos_ativos,
                (SELECT COUNT(*) FROM contrato WHERE empresa_id=emp.id) total_contratos
                FROM empresa emp WHERE emp.id=%s""", (id,), one=True)
    if not emp:
        abort(404)
    supervisores = _q("SELECT * FROM empresa_supervisor WHERE empresa_id=%s ORDER BY ordem,id", (id,))
    contatos = _q("""SELECT rc.*, u.nome as autor FROM relacionamento_contato rc
                     LEFT JOIN usuario u ON u.id=rc.usuario_id
                     WHERE rc.entidade_tipo='empresa' AND rc.entidade_id=%s
                     ORDER BY rc.created_at DESC""", (id,))
    vagas = _q("""SELECT v.*, a.nome as area_nome FROM vaga v
                  LEFT JOIN area_estagio a ON a.id=v.area_id
                  WHERE v.empresa_id=%s AND v.status IN ('aberta','em_selecao')
                  ORDER BY v.created_at DESC""", (id,))
    return render_template('empresas/detalhe.html', emp=emp, supervisores=supervisores,
                           contatos=contatos, vagas=vagas,
                           tipos_contato=TIPOS_INTERACAO)


@app.route('/empresas/<int:id>/contato', methods=['POST'])
@login_required
def empresa_contato_novo(id):
    tipo = request.form.get('tipo', '').strip()
    descricao = request.form.get('descricao', '').strip()
    if tipo and descricao:
        _run("""INSERT INTO relacionamento_contato (entidade_tipo,entidade_id,tipo,descricao,usuario_id)
                VALUES ('empresa',%s,%s,%s,%s)""", (id, tipo, descricao, current_user.id))
        flash('Contato registrado.', 'success')
    return redirect(url_for('empresa_detalhe', id=id))


@app.route('/empresas/<int:id>/nps', methods=['POST'])
@login_required
def empresa_nps_set(id):
    nps = request.form.get('nps')
    if nps and nps.isdigit() and 1 <= int(nps) <= 5:
        _run("UPDATE empresa SET nps=%s WHERE id=%s", (int(nps), id))
        flash('NPS atualizado.', 'success')
    return redirect(url_for('empresa_detalhe', id=id))


@app.route('/empresas/<int:id>/convenio')
@login_required
def empresa_convenio(id):
    emp = _q("SELECT * FROM empresa WHERE id = %s", (id,), one=True)
    if not emp:
        abort(404)
    prazo = request.args.get('prazo', 'determinado')
    data_fim_str = request.args.get('data_fim', '')
    data_fim = None
    if prazo == 'determinado' and data_fim_str:
        try:
            from datetime import datetime as _dt
            data_fim = _dt.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    cfg = _get_config()
    return render_template('empresas/convenio.html', emp=emp, prazo=prazo,
                           data_fim=data_fim, cfg=cfg, hoje=date.today())


@app.route('/api/ie_professores/<int:ie_id>')
@login_required
def api_ie_professores(ie_id):
    profs = _q("SELECT id, nome, cargo FROM ie_professor WHERE ie_id = %s ORDER BY ordem, id", (ie_id,))
    return jsonify([dict(p) for p in profs])


@app.route('/api/empresa/<int:id>')
@login_required
def api_empresa(id):
    row = _q("SELECT * FROM empresa WHERE id = %s", (id,), one=True)
    return jsonify(dict(row) if row else {})


@app.route('/api/empresa_supervisores/<int:empresa_id>')
@login_required
def api_empresa_supervisores(empresa_id):
    sups = _q("SELECT id, nome, cargo, registro FROM empresa_supervisor WHERE empresa_id = %s ORDER BY ordem, id", (empresa_id,))
    return jsonify([dict(s) for s in sups])


# ─── IEs ──────────────────────────────────────────────────────────────────────

@app.route('/ies')
@login_required
def ies():
    q = request.args.get('q', '').strip()
    if q:
        like = f'%{q}%'
        rows = _q("""SELECT ie.*,
                     (SELECT COUNT(*) FROM contrato WHERE ie_id = ie.id) qtd_contratos
                     FROM ie
                     WHERE ie.nome ILIKE %s OR ie.sigla ILIKE %s
                        OR ie.cidade ILIKE %s OR ie.cnpj ILIKE %s
                     ORDER BY ie.nome""", (like, like, like, like))
    else:
        rows = _q("""SELECT ie.*,
                     (SELECT COUNT(*) FROM contrato WHERE ie_id = ie.id) qtd_contratos
                     FROM ie ORDER BY ie.nome""")
    return render_template('ies/lista.html', ies=rows, q=q)


@app.route('/ies/nova', methods=['GET', 'POST'])
@login_required
def ie_nova():
    if request.method == 'POST':
        ie_id = _ins("""INSERT INTO ie
                (nome,sigla,cnpj,endereco,cidade,estado,telefone,email,coordenador,coordenador_cargo,
                 representante_legal,cargo_representante_legal,signatario_tce,data_vencimento_convenio)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (request.form['nome'], request.form.get('sigla'),
              request.form.get('cnpj') or None,
              request.form.get('endereco'), request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('telefone'), request.form.get('email'),
              request.form.get('coordenador'), request.form.get('coordenador_cargo'),
              request.form.get('representante_legal'), request.form.get('cargo_representante_legal'),
              request.form.get('signatario_tce', 'coordenador'),
              request.form.get('data_vencimento_convenio') or None))
        nomes_p = request.form.getlist('prof_nome[]')
        cargos_p = request.form.getlist('prof_cargo[]')
        for i, np in enumerate(nomes_p):
            cp = cargos_p[i] if i < len(cargos_p) else ''
            if np.strip():
                _ins("INSERT INTO ie_professor (ie_id,nome,cargo,ordem) VALUES (%s,%s,%s,%s)",
                     (ie_id, np.strip(), cp.strip(), i))
        _log('criar', 'ie', ie_id, f'Criou instituição: {request.form["nome"]}')
        flash('Instituição cadastrada!', 'success')
        return redirect(url_for('ies'))
    return render_template('ies/form.html', ie=None, professores=[])


@app.route('/ies/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def ie_editar(id):
    ie = _q("SELECT * FROM ie WHERE id = %s", (id,), one=True)
    if not ie:
        abort(404)
    if request.method == 'POST':
        _run("""UPDATE ie SET nome=%s,sigla=%s,cnpj=%s,endereco=%s,cidade=%s,estado=%s,telefone=%s,email=%s,
                coordenador=%s,coordenador_cargo=%s,representante_legal=%s,
                cargo_representante_legal=%s,signatario_tce=%s,
                data_vencimento_convenio=%s WHERE id=%s""",
             (request.form['nome'], request.form.get('sigla'),
              request.form.get('cnpj') or None,
              request.form.get('endereco'), request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('telefone'), request.form.get('email'),
              request.form.get('coordenador'), request.form.get('coordenador_cargo'),
              request.form.get('representante_legal'), request.form.get('cargo_representante_legal'),
              request.form.get('signatario_tce', 'coordenador'),
              request.form.get('data_vencimento_convenio') or None, id))
        _run("DELETE FROM ie_professor WHERE ie_id = %s", (id,))
        nomes_p = request.form.getlist('prof_nome[]')
        cargos_p = request.form.getlist('prof_cargo[]')
        for i, np in enumerate(nomes_p):
            cp = cargos_p[i] if i < len(cargos_p) else ''
            if np.strip():
                _ins("INSERT INTO ie_professor (ie_id,nome,cargo,ordem) VALUES (%s,%s,%s,%s)",
                     (id, np.strip(), cp.strip(), i))
        _log('editar', 'ie', id, f'Editou instituição: {request.form["nome"]}')
        flash('Atualizada!', 'success')
        return redirect(url_for('ies'))
    professores = _q("SELECT * FROM ie_professor WHERE ie_id = %s ORDER BY ordem, id", (id,))
    return render_template('ies/form.html', ie=ie, professores=professores)


@app.route('/ies/<int:id>/excluir')
@login_required
def ie_excluir(id):
    reg = _q("SELECT nome FROM ie WHERE id = %s", (id,), one=True)
    _run("DELETE FROM ie WHERE id = %s", (id,))
    _log('excluir', 'ie', id, f'Excluiu instituição: {reg["nome"] if reg else id}')
    flash('Excluída.', 'warning')
    return redirect(url_for('ies'))


@app.route('/ies/<int:id>')
@login_required
def ie_detalhe(id):
    ie = _q("""SELECT ie.*,
               (SELECT COUNT(*) FROM contrato WHERE ie_id=ie.id AND data_encerramento IS NULL) contratos_ativos,
               (SELECT COUNT(*) FROM contrato WHERE ie_id=ie.id) total_contratos
               FROM ie WHERE ie.id=%s""", (id,), one=True)
    if not ie:
        abort(404)
    professores = _q("SELECT * FROM ie_professor WHERE ie_id=%s ORDER BY ordem,id", (id,))
    contatos = _q("""SELECT rc.*, u.nome as autor FROM relacionamento_contato rc
                     LEFT JOIN usuario u ON u.id=rc.usuario_id
                     WHERE rc.entidade_tipo='ie' AND rc.entidade_id=%s
                     ORDER BY rc.created_at DESC""", (id,))
    dias_venc = None
    if ie.get('data_vencimento_convenio'):
        dias_venc = (ie['data_vencimento_convenio'] - date.today()).days
    return render_template('ies/detalhe.html', ie=ie, professores=professores,
                           contatos=contatos, dias_venc=dias_venc,
                           tipos_contato=TIPOS_INTERACAO)


@app.route('/ies/<int:id>/contato', methods=['POST'])
@login_required
def ie_contato_novo(id):
    tipo = request.form.get('tipo', '').strip()
    descricao = request.form.get('descricao', '').strip()
    if tipo and descricao:
        _run("""INSERT INTO relacionamento_contato (entidade_tipo,entidade_id,tipo,descricao,usuario_id)
                VALUES ('ie',%s,%s,%s,%s)""", (id, tipo, descricao, current_user.id))
        flash('Contato registrado.', 'success')
    return redirect(url_for('ie_detalhe', id=id))


@app.route('/ies/<int:id>/convenio')
@login_required
def ie_convenio(id):
    ie = _q("SELECT * FROM ie WHERE id = %s", (id,), one=True)
    if not ie:
        abort(404)
    prazo = request.args.get('prazo', 'determinado')
    data_fim_str = request.args.get('data_fim', '')
    data_fim = None
    if prazo == 'determinado' and data_fim_str:
        try:
            from datetime import datetime as _dt
            data_fim = _dt.strptime(data_fim_str, '%Y-%m-%d').date()
        except ValueError:
            pass
    cfg = _get_config()
    return render_template('ies/convenio.html', ie=ie, prazo=prazo,
                           data_fim=data_fim, cfg=cfg, agente=AGENTE, hoje=date.today())


# ─── ÁREAS DE ESTÁGIO ────────────────────────────────────────────────────────

@app.route('/areas')
@login_required
def areas():
    areas_list = _q("""
        SELECT a.*, COUNT(av.id) qtd_atividades
        FROM area_estagio a
        LEFT JOIN area_atividade av ON av.area_id = a.id
        GROUP BY a.id ORDER BY a.nome
    """)
    return render_template('areas/lista.html', areas=areas_list)


@app.route('/areas/nova', methods=['GET', 'POST'])
@login_required
def area_nova():
    if request.method == 'POST':
        nome = request.form['nome'].strip()
        if not nome:
            flash('Nome obrigatório.', 'danger')
            return redirect(url_for('area_nova'))
        area_id = _ins("INSERT INTO area_estagio (nome) VALUES (%s)", (nome,))
        for i, desc in enumerate(request.form.getlist('atividade[]')):
            desc = desc.strip()
            if desc:
                _run("INSERT INTO area_atividade (area_id, descricao, ordem) VALUES (%s,%s,%s)",
                     (area_id, desc, i))
        _log('criar', 'area_estagio', area_id, f'Criou área: {nome}')
        flash('Área criada!', 'success')
        return redirect(url_for('areas'))
    return render_template('areas/form.html', area=None, atividades=[])


@app.route('/areas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def area_editar(id):
    area = _q("SELECT * FROM area_estagio WHERE id = %s", (id,), one=True)
    if not area:
        abort(404)
    if request.method == 'POST':
        nome = request.form['nome'].strip()
        _run("UPDATE area_estagio SET nome=%s WHERE id=%s", (nome, id))
        _run("DELETE FROM area_atividade WHERE area_id=%s", (id,))
        for i, desc in enumerate(request.form.getlist('atividade[]')):
            desc = desc.strip()
            if desc:
                _run("INSERT INTO area_atividade (area_id, descricao, ordem) VALUES (%s,%s,%s)",
                     (id, desc, i))
        _log('editar', 'area_estagio', id, f'Editou área: {nome}')
        flash('Área atualizada!', 'success')
        return redirect(url_for('areas'))
    atividades = _q("SELECT * FROM area_atividade WHERE area_id=%s ORDER BY ordem, id", (id,))
    return render_template('areas/form.html', area=area, atividades=atividades)


@app.route('/areas/<int:id>/excluir')
@login_required
def area_excluir(id):
    reg = _q("SELECT nome FROM area_estagio WHERE id=%s", (id,), one=True)
    _run("DELETE FROM area_estagio WHERE id=%s", (id,))
    _log('excluir', 'area_estagio', id, f'Excluiu área: {reg["nome"] if reg else id}')
    flash('Área excluída.', 'warning')
    return redirect(url_for('areas'))


@app.route('/api/area_atividades/<int:area_id>')
@login_required
def api_area_atividades(area_id):
    atividades = _q(
        "SELECT descricao FROM area_atividade WHERE area_id=%s ORDER BY ordem, id",
        (area_id,))
    return jsonify([r['descricao'] for r in atividades])


@app.route('/api/ia/atividades')
@login_required
def api_ia_atividades():
    import os
    key = os.environ.get('ANTHROPIC_API_KEY', '')
    if not key:
        return jsonify({'erro': 'ANTHROPIC_API_KEY não configurada no servidor.'}), 503
    curso = request.args.get('curso', '').strip()
    area  = request.args.get('area', '').strip()
    if not curso and not area:
        return jsonify({'erro': 'Informe o curso ou a área de atuação.'}), 400
    try:
        import anthropic as _ant
        client = _ant.Anthropic(api_key=key)
        desc = f"curso de {curso}" if curso else ""
        if area:
            desc += (" na área de " if desc else "") + area
        msg = client.messages.create(
            model='claude-haiku-4-5-20251001',
            max_tokens=512,
            messages=[{
                'role': 'user',
                'content': (
                    f"Liste exatamente 8 atividades típicas de estágio para estudante de {desc}. "
                    "Cada atividade deve começar com verbo no infinitivo (ex: Auxiliar, Elaborar, Analisar). "
                    "Máximo 110 caracteres por item. Sem numeração, sem traço, sem ponto final. "
                    "Uma atividade por linha. Somente a lista, sem texto adicional."
                )
            }]
        )
        atividades = [l.strip(' -–•') for l in msg.content[0].text.strip().splitlines() if l.strip()][:8]
        return jsonify({'atividades': atividades})
    except Exception as ex:
        return jsonify({'erro': str(ex)}), 500


# ─── CONTRATOS ────────────────────────────────────────────────────────────────

@app.route('/contratos')
@login_required
def contratos():
    q = request.args.get('q', '')
    status = request.args.get('status', '')
    sql = """SELECT c.*, e.nome est_nome, emp.nome emp_nome, emp.nome_fantasia emp_nome_fantasia,
             ie.sigla ie_sigla, ie.nome ie_nome,
             COALESCE(
                 c.data_encerramento,
                 (SELECT nova_data_fim FROM aditivo
                  WHERE contrato_id = c.id AND nova_data_fim IS NOT NULL AND nova_data_fim != ''
                  ORDER BY created_at DESC LIMIT 1),
                 c.data_fim
             ) as effective_data_fim
             FROM contrato c
             JOIN estagiario e ON e.id = c.estagiario_id
             JOIN empresa emp ON emp.id = c.empresa_id
             JOIN ie ON ie.id = c.ie_id"""
    params = []
    if q:
        sql += """ WHERE (unaccent(e.nome) ILIKE unaccent(%s)
                       OR unaccent(emp.nome) ILIKE unaccent(%s)
                       OR unaccent(COALESCE(emp.nome_fantasia, emp.nome)) ILIKE unaccent(%s)
                       OR unaccent(ie.nome) ILIKE unaccent(%s)
                       OR unaccent(COALESCE(ie.sigla, '')) ILIKE unaccent(%s))"""
        params += [f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%', f'%{q}%']
    sql += ' ORDER BY effective_data_fim'
    rows = _q(sql, params)
    if status:
        filtered = []
        for r in rows:
            st = calcular_status(r['effective_data_fim'])
            enc = bool(r.get('data_encerramento'))
            if status == 'ativo' and st == 'ATIVO' and not enc:
                filtered.append(r)
            elif status == 'vencendo' and 'DIAS' in st and not enc:
                filtered.append(r)
            elif status == 'vencido' and st == 'VENCIDO' and not enc:
                filtered.append(r)
            elif status == 'encerrado' and enc:
                filtered.append(r)
        rows = filtered
    return render_template('contratos/lista.html', contratos=rows, q=q, status=status)


@app.route('/contratos/novo', methods=['GET', 'POST'])
@login_required
def contrato_novo():
    estagiarios = _q("SELECT * FROM estagiario WHERE status='ativo' ORDER BY nome")
    empresas_list = _q("SELECT id, nome, COALESCE(nome_fantasia, nome) AS display FROM empresa WHERE status='ativo' ORDER BY COALESCE(nome_fantasia, nome)")
    ies_list = _q("SELECT * FROM ie ORDER BY COALESCE(NULLIF(TRIM(sigla),''), nome)")
    areas_list = _q("SELECT id, nome FROM area_estagio WHERE status='ativo' ORDER BY nome")
    if request.method == 'POST':
        erros = []
        if not request.form.get('estagiario_id'): erros.append('Estagiário obrigatório.')
        if not request.form.get('empresa_id'):    erros.append('Empresa obrigatória.')
        if not request.form.get('ie_id'):         erros.append('Instituição de Ensino obrigatória.')
        if not request.form.get('data_inicio'):   erros.append('Data de início obrigatória.')
        if not request.form.get('data_fim'):      erros.append('Data de fim obrigatória.')
        if not request.form.get('curso'):         erros.append('Curso obrigatório.')
        if erros:
            for e in erros: flash(e, 'danger')
            return render_template('contratos/form.html', c=request.form,
                                   estagiarios=estagiarios, empresas=empresas_list,
                                   ies=ies_list, areas=areas_list, aditivos=[])
        try:
            ats = '||'.join(request.form.get(f'atividade_{i}', '') for i in range(1, 10))
            _ins("""INSERT INTO contrato
                    (estagiario_id,empresa_id,ie_id,orientador,
                     supervisor_nome,supervisor_cargo,supervisor_registro,
                     curso,tipo_estagio,area_atuacao,ch_diaria,ch_semanal,
                     data_inicio,data_fim,numero_contrato,bolsa,bolsa_tipo,taxa,aux_transporte,
                     atividades,obs,jornada,data_encerramento,ie_professor_id)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                 (request.form['estagiario_id'], request.form['empresa_id'], request.form['ie_id'],
                  request.form.get('orientador', 'Salmo Lima Costa'),
                  request.form.get('supervisor_nome'), request.form.get('supervisor_cargo'),
                  request.form.get('supervisor_registro'),
                  request.form['curso'], request.form.get('tipo_estagio', 'Não Obrigatório'),
                  request.form.get('area_atuacao'),
                  request.form.get('ch_diaria', 6), request.form.get('ch_semanal', 30),
                  request.form['data_inicio'], request.form['data_fim'],
                  request.form.get('numero_contrato'),
                  request.form.get('bolsa') or None,
                  request.form.get('bolsa_tipo', 'mensal'),
                  request.form.get('taxa') or None,
                  request.form.get('aux_transporte') or None,
                  ats, request.form.get('obs'), _build_jornada_json(),
                  request.form.get('data_encerramento') or None,
                  request.form.get('ie_professor_id') or None))
            _est = _q("SELECT nome FROM estagiario WHERE id=%s", (request.form['estagiario_id'],), one=True)
            _emp = _q("SELECT nome FROM empresa WHERE id=%s", (request.form['empresa_id'],), one=True)
            _log('criar', 'contrato', None,
                 f'Criou contrato: {_est["nome"] if _est else "?"} @ {_emp["nome"] if _emp else "?"}'
                 f' ({request.form["data_inicio"]} a {request.form["data_fim"]})')
            flash('Contrato criado!', 'success')
            return redirect(url_for('contratos'))
        except Exception as ex:
            flash(f'Erro ao salvar contrato: {ex}', 'danger')
            return render_template('contratos/form.html', c=request.form,
                                   estagiarios=estagiarios, empresas=empresas_list,
                                   ies=ies_list, areas=areas_list, aditivos=[])
    return render_template('contratos/form.html', c=None,
                           estagiarios=estagiarios, empresas=empresas_list, ies=ies_list,
                           areas=areas_list, aditivos=[])


@app.route('/contratos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def contrato_editar(id):
    c = _q("SELECT * FROM contrato WHERE id = %s", (id,), one=True)
    if not c:
        abort(404)
    estagiarios = _q("SELECT * FROM estagiario WHERE status='ativo' ORDER BY nome")
    empresas_list = _q("SELECT id, nome, COALESCE(nome_fantasia, nome) AS display FROM empresa WHERE status='ativo' ORDER BY COALESCE(nome_fantasia, nome)")
    ies_list = _q("SELECT * FROM ie ORDER BY COALESCE(NULLIF(TRIM(sigla),''), nome)")
    areas_list = _q("SELECT id, nome FROM area_estagio WHERE status='ativo' ORDER BY nome")
    aditivos = _q("SELECT * FROM aditivo WHERE contrato_id = %s ORDER BY created_at", (id,))
    relatorios = _q("SELECT * FROM relatorio_periodo WHERE contrato_id=%s ORDER BY numero", (id,))
    if request.method == 'POST':
        erros = []
        if not request.form.get('estagiario_id'): erros.append('Estagiário obrigatório.')
        if not request.form.get('empresa_id'):    erros.append('Empresa obrigatória.')
        if not request.form.get('ie_id'):         erros.append('Instituição de Ensino obrigatória.')
        if not request.form.get('data_inicio'):   erros.append('Data de início obrigatória.')
        if not request.form.get('data_fim'):      erros.append('Data de fim obrigatória.')
        if not request.form.get('curso'):         erros.append('Curso obrigatório.')
        if erros:
            for e in erros: flash(e, 'danger')
            return render_template('contratos/form.html', c=request.form,
                                   estagiarios=estagiarios, empresas=empresas_list,
                                   ies=ies_list, areas=areas_list, aditivos=aditivos,
                                   relatorios=relatorios, relatorio_pendente=False)
        try:
            ats = '||'.join(request.form.get(f'atividade_{i}', '') for i in range(1, 10))
            _run("""UPDATE contrato SET
                    estagiario_id=%s,empresa_id=%s,ie_id=%s,orientador=%s,
                    supervisor_nome=%s,supervisor_cargo=%s,supervisor_registro=%s,
                    curso=%s,tipo_estagio=%s,area_atuacao=%s,ch_diaria=%s,ch_semanal=%s,
                    data_inicio=%s,data_fim=%s,numero_contrato=%s,bolsa=%s,bolsa_tipo=%s,
                    taxa=%s,aux_transporte=%s,atividades=%s,obs=%s,jornada=%s,
                    data_encerramento=%s,ie_professor_id=%s WHERE id=%s""",
                 (request.form['estagiario_id'], request.form['empresa_id'], request.form['ie_id'],
                  request.form.get('orientador'),
                  request.form.get('supervisor_nome'), request.form.get('supervisor_cargo'),
                  request.form.get('supervisor_registro'),
                  request.form['curso'], request.form.get('tipo_estagio'),
                  request.form.get('area_atuacao'),
                  request.form.get('ch_diaria'), request.form.get('ch_semanal'),
                  request.form['data_inicio'], request.form['data_fim'],
                  request.form.get('numero_contrato'),
                  request.form.get('bolsa') or None,
                  request.form.get('bolsa_tipo', 'mensal'),
                  request.form.get('taxa') or None,
                  request.form.get('aux_transporte') or None,
                  ats, request.form.get('obs'), _build_jornada_json(),
                  request.form.get('data_encerramento') or None,
                  request.form.get('ie_professor_id') or None, id))
            _est2 = _q("SELECT nome FROM estagiario WHERE id=%s", (request.form['estagiario_id'],), one=True)
            _emp2 = _q("SELECT nome FROM empresa WHERE id=%s", (request.form['empresa_id'],), one=True)
            _log('editar', 'contrato', id,
                 f'Editou contrato: {_est2["nome"] if _est2 else "?"} @ {_emp2["nome"] if _emp2 else "?"}')
            flash('Contrato atualizado!', 'success')
            return redirect(url_for('contratos'))
        except Exception as ex:
            flash(f'Erro ao salvar contrato: {ex}', 'danger')
            return render_template('contratos/form.html', c=request.form,
                                   estagiarios=estagiarios, empresas=empresas_list,
                                   ies=ies_list, areas=areas_list, aditivos=aditivos,
                                   relatorios=relatorios, relatorio_pendente=False)
    last_rel = relatorios[-1] if relatorios else None
    if last_rel:
        prox_inicio = date.fromisoformat(str(last_rel['data_fim'])[:10]) + timedelta(days=1)
    else:
        try:
            prox_inicio = date.fromisoformat(str(c['data_inicio'])[:10])
        except Exception:
            prox_inicio = date.today()
    relatorio_pendente = (date.today() - prox_inicio).days >= 180
    return render_template('contratos/form.html', c=c,
                           estagiarios=estagiarios, empresas=empresas_list, ies=ies_list,
                           areas=areas_list, aditivos=aditivos,
                           relatorios=relatorios, relatorio_pendente=relatorio_pendente)


@app.route('/contratos/<int:id>/excluir')
@login_required
def contrato_excluir(id):
    info = _q("""SELECT e.nome est, emp.nome emp FROM contrato c
                 JOIN estagiario e ON e.id = c.estagiario_id
                 JOIN empresa emp ON emp.id = c.empresa_id WHERE c.id = %s""", (id,), one=True)
    _run("DELETE FROM contrato WHERE id = %s", (id,))
    desc = f'Excluiu contrato: {info["est"]} @ {info["emp"]}' if info else f'Excluiu contrato ID {id}'
    _log('excluir', 'contrato', id, desc)
    flash('Excluído.', 'warning')
    return redirect(url_for('contratos'))


# ─── DOCUMENTOS ───────────────────────────────────────────────────────────────

def _doc_ctx(id):
    c = _q("SELECT * FROM contrato WHERE id = %s", (id,), one=True)
    if not c:
        return None
    est = _q("SELECT * FROM estagiario WHERE id = %s", (c['estagiario_id'],), one=True)
    emp = _q("SELECT * FROM empresa WHERE id = %s", (c['empresa_id'],), one=True)
    ie = _q("SELECT * FROM ie WHERE id = %s", (c['ie_id'],), one=True)
    cfg = _get_config()
    aditivos = _q("SELECT * FROM aditivo WHERE contrato_id = %s ORDER BY created_at", (id,))

    try:
        ini = datetime.strptime(str(c['data_inicio'])[:10], '%Y-%m-%d').date()
        fim = datetime.strptime(str(c['data_fim'])[:10], '%Y-%m-%d').date()
        meses = (fim.year - ini.year) * 12 + fim.month - ini.month + 1
        ch_total = meses * 4 * (c['ch_semanal'] or 30)
    except Exception:
        ch_total = 0

    enc_str = c.get('data_encerramento')
    try:
        if enc_str:
            enc = datetime.strptime(str(enc_str)[:10], '%Y-%m-%d').date()
            meses_real = (enc.year - ini.year) * 12 + enc.month - ini.month + 1
            ch_total_real = meses_real * 4 * (c['ch_semanal'] or 30)
        else:
            ch_total_real = ch_total
    except Exception:
        ch_total_real = ch_total

    ch_j_diaria, ch_j_semanal = calcular_ch_jornada(c.get('jornada'))

    professor = None
    if c.get('ie_professor_id'):
        professor = _q("SELECT * FROM ie_professor WHERE id = %s", (c['ie_professor_id'],), one=True)

    sig_tce = (ie.get('signatario_tce') or 'coordenador') if ie else 'coordenador'
    if sig_tce == 'representante':
        sig_nome = ie.get('representante_legal', '') if ie else ''
        sig_cargo = ie.get('cargo_representante_legal', '') if ie else ''
        sig_tipo = 'Representante Legal'
    elif sig_tce == 'professor':
        sig_nome = professor['nome'] if professor else ''
        sig_cargo = professor['cargo'] if professor else ''
        sig_tipo = 'Professor Orientador'
    else:
        sig_nome = ie.get('coordenador', '') if ie else ''
        sig_cargo = ie.get('coordenador_cargo', '') if ie else ''
        sig_tipo = 'Coordenador(a) de Estágio'

    d = type('D', (), {
        'id': c['id'],
        'curso': c['curso'],
        'tipo_estagio': c['tipo_estagio'],
        'area_atuacao': c['area_atuacao'] or c['curso'],
        'ch_diaria': c['ch_diaria'] or 6,
        'ch_semanal': c['ch_semanal'] or 30,
        'data_inicio': c['data_inicio'],
        'data_fim': c['data_fim'],
        'data_encerramento': enc_str,
        'numero_contrato': c['numero_contrato'],
        'bolsa': c['bolsa'],
        'bolsa_tipo': c.get('bolsa_tipo') or 'mensal',
        'bolsa_extenso': num_extenso(c['bolsa']) if c['bolsa'] else '',
        'taxa': c['taxa'],
        'aux_transporte': c['aux_transporte'],
        'aux_transporte_extenso': num_extenso(c['aux_transporte']) if c['aux_transporte'] else '',
        'atividades': c['atividades'],
        'orientador': c['orientador'],
        'supervisor_nome': c['supervisor_nome'],
        'supervisor_cargo': c['supervisor_cargo'],
        'supervisor_registro': c['supervisor_registro'],
        'num_relatorio': c['num_relatorio'] or 1,
        'jornada_texto': formatar_jornada(c['jornada']),
        'ch_diaria_jornada': ch_j_diaria,
        'ch_semanal_jornada': ch_j_semanal,
        'seg_seguradora': cfg.get('seg_seguradora', ''),
        'seg_apolice': cfg.get('seg_apolice', ''),
        'seg_coberturas': cfg.get('seg_coberturas', ''),
        'seg_vigencia': cfg.get('seg_vigencia', ''),
        'est_nome': est['nome'] if est else '',
        'est_cpf': est['cpf'] if est else '',
        'est_rg': est['rg'] if est else '',
        'est_data_nasc': est['data_nascimento'] if est else '',
        'est_telefone': est['telefone'] if est else '',
        'est_email': est['email'] if est else '',
        'est_endereco': est['endereco'] if est else '',
        'est_semestre': _fmt_semestre(est['semestre'], est['tipo_ensino']) if est else '',
        'est_matricula': est['matricula'] if est else '',
        'emp_nome': emp['nome'] if emp else '',
        'emp_cnpj': emp['cnpj'] if emp else '',
        'emp_endereco': emp['endereco'] if emp else '',
        'emp_cidade': emp['cidade'] if emp else 'Vitória da Conquista',
        'emp_telefone': emp['telefone'] if emp else '',
        'emp_email': emp['email'] if emp else '',
        'emp_representante': emp['representante'] if emp else '',
        'emp_cargo_rep': emp['cargo_representante'] if emp else '',
        'emp_cpf_rep': emp.get('cpf_representante', '') if emp else '',
        'ie_nome': ie['nome'] if ie else '',
        'ie_sigla': ie['sigla'] if ie else '',
        'ie_cnpj': ie.get('cnpj', '') if ie else '',
        'ie_endereco': ie['endereco'] if ie else '',
        'ie_cidade': ie['cidade'] if ie else 'Vitória da Conquista',
        'ie_coordenador': ie['coordenador'] if ie else '',
        'ie_coord_cargo': ie['coordenador_cargo'] if ie else '',
        'ie_representante_legal': ie.get('representante_legal', '') if ie else '',
        'ie_cargo_rep_legal': ie.get('cargo_representante_legal', '') if ie else '',
        'ie_professor_nome': professor['nome'] if professor else '',
        'ie_professor_cargo': professor['cargo'] if professor else '',
        'ie_signatario_nome': sig_nome,
        'ie_signatario_cargo': sig_cargo,
        'ie_signatario_tipo': sig_tipo,
    })()

    return dict(d=d, agente=AGENTE, ch_total=ch_total, ch_total_real=ch_total_real,
                data_hoje=fmt_date(date.today()), fmt_date=fmt_date,
                br_currency=br_currency, aditivos=aditivos)


@app.route('/contratos/<int:id>/tce')
@login_required
def doc_tce(id):
    ctx = _doc_ctx(id)
    if not ctx:
        flash('Contrato não encontrado.', 'danger')
        return redirect(url_for('contratos'))
    return render_template('docs/tce.html', **ctx)


@app.route('/contratos/<int:id>/plano')
@login_required
def doc_plano(id):
    ctx = _doc_ctx(id)
    if not ctx:
        flash('Contrato não encontrado.', 'danger')
        return redirect(url_for('contratos'))
    return render_template('docs/plano.html', **ctx)


@app.route('/contratos/<int:id>/ciencia')
@login_required
def doc_ciencia(id):
    ctx = _doc_ctx(id)
    if not ctx:
        flash('Contrato não encontrado.', 'danger')
        return redirect(url_for('contratos'))
    return render_template('docs/ciencia.html', **ctx)


@app.route('/contratos/<int:id>/tre')
@login_required
def doc_tre(id):
    ctx = _doc_ctx(id)
    if not ctx:
        flash('Contrato não encontrado.', 'danger')
        return redirect(url_for('contratos'))
    tre_enc = request.args.get('data_encerramento', '').strip()
    if tre_enc:
        ctx['d'].data_encerramento = tre_enc
        try:
            ini = datetime.strptime(str(ctx['d'].data_inicio)[:10], '%Y-%m-%d').date()
            enc = datetime.strptime(tre_enc[:10], '%Y-%m-%d').date()
            meses_real = (enc.year - ini.year) * 12 + enc.month - ini.month + 1
            ctx['ch_total_real'] = meses_real * 4 * (ctx['d'].ch_semanal or 30)
        except Exception:
            pass
    return render_template('docs/tre.html', **ctx)


@app.route('/contratos/<int:id>/relatorio')
@login_required
def doc_relatorio(id):
    ctx = _doc_ctx(id)
    if not ctx:
        flash('Contrato não encontrado.', 'danger')
        return redirect(url_for('contratos'))
    ctx['rel_data_inicio'] = request.args.get('data_inicio', '')
    ctx['rel_data_fim'] = request.args.get('data_fim', '')
    ctx['rel_numero'] = request.args.get('num', '')
    return render_template('docs/relatorio.html', **ctx)


@app.route('/contratos/<int:id>/relatorio/proximo')
@login_required
def relatorio_proximo(id):
    ct = _q("SELECT data_inicio FROM contrato WHERE id=%s", (id,), one=True)
    if not ct:
        return jsonify(ok=False), 404
    last = _q("SELECT numero, data_fim FROM relatorio_periodo WHERE contrato_id=%s ORDER BY numero DESC LIMIT 1", (id,), one=True)
    if last:
        prox_inicio = (date.fromisoformat(str(last['data_fim'])[:10]) + timedelta(days=1)).isoformat()
        prox_num = last['numero'] + 1
    else:
        prox_inicio = str(ct['data_inicio'])[:10]
        prox_num = 1
    return jsonify(ok=True, data_inicio=prox_inicio, proximo_numero=prox_num)


@app.route('/contratos/<int:id>/relatorio/registrar', methods=['POST'])
@login_required
def relatorio_registrar(id):
    data = request.get_json()
    data_fim_str = (data or {}).get('data_fim')
    if not data_fim_str:
        return jsonify(ok=False, erro='data_fim obrigatório'), 400
    ct = _q("SELECT data_inicio FROM contrato WHERE id=%s", (id,), one=True)
    if not ct:
        return jsonify(ok=False, erro='Contrato não encontrado'), 404
    last = _q("SELECT numero, data_fim FROM relatorio_periodo WHERE contrato_id=%s ORDER BY numero DESC LIMIT 1", (id,), one=True)
    if last:
        inicio = date.fromisoformat(str(last['data_fim'])[:10]) + timedelta(days=1)
        num = last['numero']
    else:
        inicio = date.fromisoformat(str(ct['data_inicio'])[:10])
        num = 0
    fim = date.fromisoformat(data_fim_str)
    if fim < inicio:
        return jsonify(ok=False, erro='Data final anterior ao início do próximo período'), 400
    periodos = _split_periodos_6meses(inicio, fim)
    result = []
    for (ini_dt, fim_dt) in periodos:
        num += 1
        _run("INSERT INTO relatorio_periodo (contrato_id, numero, data_inicio, data_fim) VALUES (%s,%s,%s,%s)",
             (id, num, ini_dt.isoformat(), fim_dt.isoformat()))
        result.append({'numero': num, 'data_inicio': ini_dt.isoformat(), 'data_fim': fim_dt.isoformat()})
    return jsonify(ok=True, periodos=result)


@app.route('/contratos/<int:id>/aditivo')
@login_required
def doc_aditivo(id):
    ctx = _doc_ctx(id)
    if not ctx:
        flash('Contrato não encontrado.', 'danger')
        return redirect(url_for('contratos'))
    ctx['nova_data_fim'] = request.args.get('nova_data_fim', '')
    ctx['aditivo_num'] = request.args.get('aditivo_num', '')
    clausulas_extras = []
    i = 1
    while i <= 10:
        titulo = request.args.get(f'clausula_{i}_titulo', '').strip()
        texto = request.args.get(f'clausula_{i}_texto', '').strip()
        if not titulo and not texto:
            break
        clausulas_extras.append({'titulo': titulo, 'texto': texto})
        i += 1
    ctx['clausulas_extras'] = clausulas_extras
    return render_template('docs/aditivo.html', **ctx)


@app.route('/contratos/<int:id>/aditivo/registrar', methods=['POST'])
@login_required
def doc_aditivo_registrar(id):
    c = _q("SELECT id FROM contrato WHERE id = %s", (id,), one=True)
    if not c:
        return jsonify({'ok': False}), 404
    data = request.get_json(silent=True) or {}
    nova_data_fim = data.get('nova_data_fim') or None
    clausulas = data.get('clausulas', [])
    _ins("INSERT INTO aditivo (contrato_id, nova_data_fim, clausulas) VALUES (%s, %s, %s)",
         (id, nova_data_fim, json.dumps(clausulas, ensure_ascii=False)))
    numero = _q("SELECT COUNT(*) as n FROM aditivo WHERE contrato_id = %s", (id,), one=True)['n']
    _info_adi = _q("""SELECT e.nome est, emp.nome emp FROM contrato ct
                      JOIN estagiario e ON e.id=ct.estagiario_id
                      JOIN empresa emp ON emp.id=ct.empresa_id WHERE ct.id=%s""", (id,), one=True)
    _log('criar', 'aditivo', id,
         f'Gerou {numero}º aditivo: {_info_adi["est"]} @ {_info_adi["emp"]}' if _info_adi
         else f'Gerou {numero}º aditivo do contrato ID {id}')
    return jsonify({'ok': True, 'numero': numero})


@app.route('/contratos/<int:id>/aditivo/<int:aditivo_id>/editar', methods=['POST'])
@login_required
def aditivo_editar(id, aditivo_id):
    ad = _q("SELECT id FROM aditivo WHERE id=%s AND contrato_id=%s", (aditivo_id, id), one=True)
    if not ad:
        return jsonify({'ok': False}), 404
    data = request.get_json(silent=True) or {}
    nova_data_fim = data.get('nova_data_fim') or None
    clausulas = data.get('clausulas', [])
    _run("UPDATE aditivo SET nova_data_fim=%s, clausulas=%s WHERE id=%s",
         (nova_data_fim, json.dumps(clausulas, ensure_ascii=False), aditivo_id))
    _log('editar', 'aditivo', aditivo_id, f'Editou aditivo do contrato ID {id}')
    return jsonify({'ok': True})


@app.route('/contratos/<int:id>/aditivo/<int:aditivo_id>/excluir', methods=['POST'])
@login_required
def aditivo_excluir(id, aditivo_id):
    ad = _q("SELECT nova_data_fim FROM aditivo WHERE id=%s AND contrato_id=%s", (aditivo_id, id), one=True)
    if not ad:
        abort(404)
    _run("DELETE FROM aditivo WHERE id=%s", (aditivo_id,))
    # vigência efetiva após exclusão
    prev = _q("SELECT nova_data_fim FROM aditivo WHERE contrato_id=%s AND nova_data_fim IS NOT NULL ORDER BY created_at DESC LIMIT 1", (id,), one=True)
    if prev:
        vig = fmt_date(prev['nova_data_fim'])
    else:
        ct = _q("SELECT data_fim FROM contrato WHERE id=%s", (id,), one=True)
        vig = fmt_date(ct['data_fim']) if ct else '—'
    _log('excluir', 'aditivo', aditivo_id, f'Excluiu aditivo do contrato ID {id}')
    flash(f'Aditivo excluído. Vigência efetiva do contrato: {vig}.', 'warning')
    return redirect(url_for('contrato_editar', id=id))


@app.route('/api/check-vinculo/<int:est_id>/<int:emp_id>')
@login_required
def api_check_vinculo(est_id, emp_id):
    exclude_id = request.args.get('exclude', type=int)
    sql = "SELECT data_inicio, data_encerramento, data_fim FROM contrato WHERE estagiario_id = %s AND empresa_id = %s"
    params = [est_id, emp_id]
    if exclude_id:
        sql += " AND id != %s"
        params.append(exclude_id)
    cts = _q(sql, params)
    total_dias = 0
    for ct in cts:
        try:
            ini = datetime.strptime(str(ct['data_inicio'])[:10], '%Y-%m-%d').date()
            fim_str = ct['data_encerramento'] or ct['data_fim']
            fim = datetime.strptime(str(fim_str)[:10], '%Y-%m-%d').date()
            total_dias += (fim - ini).days
        except Exception:
            pass
    anos = round(total_dias / 365.25, 2)
    return jsonify({'total_dias': total_dias, 'anos': anos,
                    'limite_atingido': total_dias >= 730, 'aviso': total_dias >= 600})


# ─── ADMIN — CONFIGURAÇÕES ────────────────────────────────────────────────────

@app.route('/admin/config', methods=['GET', 'POST'])
@admin_required
def admin_config():
    if request.method == 'POST':
        for chave in ['seg_seguradora', 'seg_apolice', 'seg_coberturas', 'seg_vigencia']:
            valor = request.form.get(chave, '').strip()
            _run("INSERT INTO config (chave, valor) VALUES (%s, %s) ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
                 (chave, valor))
        _log('editar', 'config', None, 'Editou configurações de seguro')
        flash('Configurações salvas!', 'success')
        return redirect(url_for('admin_config'))
    cfg = _get_config()
    return render_template('admin/config.html', cfg=cfg)


# ─── RELATÓRIOS ───────────────────────────────────────────────────────────────

@app.route('/relatorio/vencimentos')
@login_required
def relatorio_vencimentos():
    data_ate = request.args.get('data_ate', '')
    base_sql = """
        SELECT c.*, e.nome est_nome, emp.nome emp_nome, ie.sigla ie_sigla,
               COALESCE(
                   c.data_encerramento,
                   (SELECT nova_data_fim FROM aditivo
                    WHERE contrato_id = c.id AND nova_data_fim IS NOT NULL AND nova_data_fim != ''
                    ORDER BY created_at DESC LIMIT 1),
                   c.data_fim
               ) as effective_data_fim
        FROM contrato c
        JOIN estagiario e ON e.id = c.estagiario_id
        JOIN empresa emp ON emp.id = c.empresa_id
        JOIN ie ON ie.id = c.ie_id
    """
    if data_ate:
        sql = f"WITH eff AS ({base_sql}) SELECT * FROM eff WHERE effective_data_fim <= %s ORDER BY effective_data_fim"
        rows = _q(sql, (data_ate,))
    else:
        sql = f"WITH eff AS ({base_sql}) SELECT * FROM eff ORDER BY effective_data_fim"
        rows = _q(sql)
    return render_template('relatorio/vencimentos.html', contratos=rows, data_ate=data_ate)


@app.route('/relatorio/estagiarios')
@login_required
def relatorio_estagiarios():
    empresa_id = request.args.get('empresa_id', '')
    ie_id = request.args.get('ie_id', '')
    curso = request.args.get('curso', '').strip()
    modo = request.args.get('modo', '')

    sql = """
        SELECT c.*, e.nome est_nome, e.cpf est_cpf, e.semestre est_semestre,
               e.tipo_ensino est_tipo_ensino, e.matricula est_matricula,
               emp.nome emp_nome, COALESCE(emp.nome_fantasia, emp.nome) emp_display,
               ie.nome ie_nome, ie.sigla ie_sigla
        FROM contrato c
        JOIN estagiario e ON e.id = c.estagiario_id
        JOIN empresa emp ON emp.id = c.empresa_id
        JOIN ie ON ie.id = c.ie_id
        WHERE c.data_encerramento IS NULL
    """
    params = []
    if empresa_id:
        sql += " AND c.empresa_id = %s"
        params.append(empresa_id)
    if ie_id:
        sql += " AND c.ie_id = %s"
        params.append(ie_id)
    if curso:
        sql += " AND c.curso ILIKE %s"
        params.append(f'%{curso}%')
    sql += " ORDER BY emp.nome, e.nome"

    contratos = _q(sql, params)
    empresas = _q("SELECT id, nome FROM empresa WHERE status='ativo' ORDER BY nome")
    ies = _q("SELECT id, nome FROM ie ORDER BY nome")
    total_taxa = sum(c['taxa'] or 0 for c in contratos)
    total_bolsa = sum(c['bolsa'] or 0 for c in contratos)
    empresa_sel = _q("SELECT nome FROM empresa WHERE id = %s", (empresa_id,), one=True) if empresa_id else None

    # Agrupar por empresa para modo subtotais
    grupos = {}
    for c in contratos:
        k = c['emp_nome']
        if k not in grupos:
            grupos[k] = {'display': c['emp_display'], 'itens': [], 'total_taxa': 0}
        grupos[k]['itens'].append(c)
        grupos[k]['total_taxa'] += (c['taxa'] or 0)

    return render_template('relatorio/estagiarios.html',
                           contratos=contratos, empresas=empresas, ies=ies,
                           empresa_id=empresa_id, ie_id=ie_id, curso=curso,
                           modo=modo,
                           total_taxa=total_taxa, total_bolsa=total_bolsa,
                           grupos=grupos,
                           empresa_sel=empresa_sel, agente=AGENTE,
                           fmt_date=fmt_date, data_hoje=date.today())


@app.route('/relatorio/lista-estagiarios')
@login_required
def relatorio_lista_estagiarios():
    estado = request.args.get('estado', '').strip()
    cidade = request.args.get('cidade', '').strip()
    tipo_ensino = request.args.get('tipo_ensino', '').strip()

    sql = "SELECT * FROM estagiario WHERE status='ativo'"
    params = []
    if estado:
        sql += " AND estado = %s"
        params.append(estado)
    if cidade:
        sql += " AND cidade = %s"
        params.append(cidade)
    if tipo_ensino:
        sql += " AND tipo_ensino = %s"
        params.append(tipo_ensino)
    sql += " ORDER BY nome"

    estagiarios = _q(sql, params)
    estados_db = _q(
        "SELECT DISTINCT estado FROM estagiario WHERE estado IS NOT NULL AND estado <> '' ORDER BY estado")
    cidades_db = _q(
        "SELECT DISTINCT cidade FROM estagiario WHERE cidade IS NOT NULL AND cidade <> '' ORDER BY cidade")

    return render_template('relatorio/lista_estagiarios.html',
                           estagiarios=estagiarios, estado=estado, cidade=cidade,
                           tipo_ensino=tipo_ensino,
                           estados_db=estados_db, cidades_db=cidades_db,
                           agente=AGENTE, fmt_date=fmt_date, data_hoje=date.today())


# ─── CADASTRO PÚBLICO ────────────────────────────────────────────────────────

@app.route('/cadastro/estagiario', methods=['GET', 'POST'])
def cadastro_estagiario():
    ies = _q("SELECT id, nome, sigla FROM ie ORDER BY nome")
    if request.method == 'POST':
        try:
            _ins("""INSERT INTO estagiario
                    (nome,cpf,rg,data_nascimento,telefone,email,endereco,
                     cidade,estado,tipo_ensino,semestre,matricula,ie_id,obs,status)
                    VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pendente')""",
                 (request.form['nome'], request.form['cpf'],
                  request.form.get('rg'), request.form.get('data_nascimento') or None,
                  request.form.get('telefone'), request.form.get('email'),
                  request.form.get('endereco'),
                  request.form.get('cidade') or None,
                  request.form.get('estado') or None,
                  request.form.get('tipo_ensino', 'superior'),
                  request.form.get('semestre') or None,
                  request.form.get('matricula') or None,
                  request.form.get('ie_id') or None,
                  request.form.get('obs')))
            return render_template('cadastro/sucesso.html', tipo='estagiário')
        except psycopg2.errors.UniqueViolation:
            flash('Este CPF já está cadastrado no sistema.', 'danger')
    return render_template('cadastro/estagiario.html', ies=ies)


@app.route('/cadastro/empresa', methods=['GET', 'POST'])
def cadastro_empresa():
    if request.method == 'POST':
        _ins("""INSERT INTO empresa
                (nome,cnpj,endereco,cidade,estado,telefone,email,ramo,
                 representante,cargo_representante,cpf_representante,supervisor_nome,supervisor_cargo,status)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'pendente')""",
             (request.form['nome'], request.form.get('cnpj'),
              request.form.get('endereco'), request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('telefone'), request.form.get('email'),
              request.form.get('ramo'), request.form.get('representante'),
              request.form.get('cargo_representante'), request.form.get('cpf_representante'),
              request.form.get('supervisor_nome'), request.form.get('supervisor_cargo')))
        return render_template('cadastro/sucesso.html', tipo='empresa')
    return render_template('cadastro/empresa.html')


# ─── ADMIN — PENDENTES ────────────────────────────────────────────────────────

@app.route('/admin/pendentes')
@admin_required
def admin_pendentes():
    estagiarios_p = _q("SELECT * FROM estagiario WHERE status='pendente' ORDER BY nome")
    empresas_p = _q("SELECT * FROM empresa WHERE status='pendente' ORDER BY nome")
    return render_template('admin/pendentes.html', estagiarios=estagiarios_p, empresas=empresas_p)


@app.route('/admin/pendentes/estagiario/<int:id>/aprovar')
@admin_required
def aprovar_estagiario(id):
    reg = _q("SELECT nome FROM estagiario WHERE id = %s", (id,), one=True)
    _run("UPDATE estagiario SET status='ativo' WHERE id=%s", (id,))
    _log('aprovar', 'estagiario', id, f'Aprovou estagiário: {reg["nome"] if reg else id}')
    flash('Estagiário aprovado com sucesso!', 'success')
    return redirect(url_for('admin_pendentes'))


@app.route('/admin/pendentes/empresa/<int:id>/aprovar')
@admin_required
def aprovar_empresa(id):
    reg = _q("SELECT nome FROM empresa WHERE id = %s", (id,), one=True)
    _run("UPDATE empresa SET status='ativo' WHERE id=%s", (id,))
    _log('aprovar', 'empresa', id, f'Aprovou empresa: {reg["nome"] if reg else id}')
    flash('Empresa aprovada com sucesso!', 'success')
    return redirect(url_for('admin_pendentes'))


@app.route('/admin/pendentes/estagiario/<int:id>/rejeitar')
@admin_required
def rejeitar_estagiario(id):
    reg = _q("SELECT nome FROM estagiario WHERE id = %s AND status='pendente'", (id,), one=True)
    _run("DELETE FROM estagiario WHERE id=%s AND status='pendente'", (id,))
    _log('rejeitar', 'estagiario', id, f'Rejeitou estagiário: {reg["nome"] if reg else id}')
    flash('Cadastro rejeitado e removido.', 'warning')
    return redirect(url_for('admin_pendentes'))


@app.route('/admin/pendentes/empresa/<int:id>/rejeitar')
@admin_required
def rejeitar_empresa(id):
    reg = _q("SELECT nome FROM empresa WHERE id = %s AND status='pendente'", (id,), one=True)
    _run("DELETE FROM empresa WHERE id=%s AND status='pendente'", (id,))
    _log('rejeitar', 'empresa', id, f'Rejeitou empresa: {reg["nome"] if reg else id}')
    flash('Cadastro rejeitado e removido.', 'warning')
    return redirect(url_for('admin_pendentes'))


# ─── ADMIN — LOG DE AUDITORIA ─────────────────────────────────────────────────

@app.route('/admin/log')
@admin_required
def admin_log():
    periodo = request.args.get('periodo', '30')
    entidade = request.args.get('entidade', '')
    uid_filtro = request.args.get('usuario_id', '')
    try:
        dias = int(periodo)
    except ValueError:
        dias = 30
    dt_from = datetime.now() - timedelta(days=dias)
    sql = """SELECT l.*, TO_CHAR(l.created_at, 'DD/MM/YYYY HH24:MI:SS') as dt_fmt
             FROM log_auditoria l WHERE l.created_at >= %s"""
    params = [dt_from]
    if entidade:
        sql += " AND l.entidade = %s"
        params.append(entidade)
    if uid_filtro:
        sql += " AND l.usuario_id = %s"
        params.append(uid_filtro)
    sql += " ORDER BY l.created_at DESC LIMIT 500"
    logs = _q(sql, params)
    usuarios = _q("SELECT id, COALESCE(nome, username) as nome FROM usuario ORDER BY nome")
    return render_template('admin/log.html', logs=logs, periodo=periodo,
                           entidade=entidade, usuario_id=uid_filtro, usuarios=usuarios)


# ─── BACKUP / RESTAURAÇÃO ─────────────────────────────────────────────────────

_BACKUP_TABELAS_EXPORT = [
    'usuario', 'estagiario', 'empresa', 'ie', 'ie_professor',
    'empresa_supervisor', 'area_estagio', 'area_atividade',
    'contrato', 'aditivo', 'config', 'log_auditoria',
]
_BACKUP_TABELAS_DELETE = list(reversed(_BACKUP_TABELAS_EXPORT))


def _get_conn_direct():
    url = DATABASE_URL
    if url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)
    return psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)


def _serial(obj):
    if isinstance(obj, (date, datetime)):
        return obj.isoformat()
    return str(obj)


def gerar_backup_json():
    conn = _get_conn_direct()
    dados = {'versao': '2.0', 'gerado_em': datetime.now().isoformat(), 'tabelas': {}}
    try:
        with conn.cursor() as cur:
            for tabela in _BACKUP_TABELAS_EXPORT:
                try:
                    cur.execute(f'SELECT * FROM {tabela}')
                    dados['tabelas'][tabela] = [dict(r) for r in cur.fetchall()]
                except Exception:
                    dados['tabelas'][tabela] = []
    finally:
        conn.close()
    return json.dumps(dados, default=_serial, ensure_ascii=False, indent=2)


def restaurar_backup_json(json_str):
    dados = json.loads(json_str)
    if dados.get('versao') not in ('1.0', '2.0'):
        raise ValueError('Versão de backup incompatível.')
    conn = _get_conn_direct()
    conn.autocommit = False
    try:
        with conn.cursor() as cur:
            cur.execute("SET session_replication_role = 'replica'")
            for tabela in _BACKUP_TABELAS_DELETE:
                cur.execute(f'DELETE FROM {tabela}')
            for tabela in _BACKUP_TABELAS_EXPORT:
                rows = dados['tabelas'].get(tabela, [])
                if not rows:
                    continue
                for row in rows:
                    cols = list(row.keys())
                    vals = list(row.values())
                    ph = ', '.join(['%s'] * len(cols))
                    cur.execute(
                        f"INSERT INTO {tabela} ({', '.join(cols)}) VALUES ({ph})",
                        vals
                    )
                cur.execute(
                    f"SELECT setval(pg_get_serial_sequence('{tabela}', 'id'), "
                    f"COALESCE(MAX(id), 1)) FROM {tabela}"
                )
            cur.execute("SET session_replication_role = 'DEFAULT'")
        conn.commit()
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


def enviar_backup_email():
    smtp_user = os.environ.get('SMTP_USER', '')
    smtp_pass = os.environ.get('SMTP_PASS', '')
    dest = os.environ.get('BACKUP_EMAIL', smtp_user)
    if not smtp_user or not smtp_pass:
        return
    try:
        json_str = gerar_backup_json()
        nome_arquivo = f"backup_ciclorh_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        msg = MIMEMultipart()
        msg['From'] = smtp_user
        msg['To'] = dest
        msg['Subject'] = f'Backup CicloRH — {datetime.now().strftime("%d/%m/%Y")}'
        msg.attach(MIMEText(
            f'Backup automático do sistema Ciclo RH gerado em '
            f'{datetime.now().strftime("%d/%m/%Y às %H:%M")}.\n\n'
            f'Tabelas incluídas: {", ".join(_BACKUP_TABELAS_EXPORT)}\n\n'
            f'Arquivo: {nome_arquivo}',
            'plain', 'utf-8'
        ))
        anexo = MIMEApplication(json_str.encode('utf-8'), Name=nome_arquivo)
        anexo['Content-Disposition'] = f'attachment; filename="{nome_arquivo}"'
        msg.attach(anexo)
        with smtplib.SMTP('smtp.gmail.com', 587) as srv:
            srv.starttls()
            srv.login(smtp_user, smtp_pass)
            srv.sendmail(smtp_user, dest, msg.as_string())
    except Exception as e:
        print(f'[Backup] Erro ao enviar e-mail: {e}')


@app.route('/admin/backup', methods=['GET', 'POST'])
@admin_required
def admin_backup():
    if request.method == 'POST':
        acao = request.form.get('acao', 'download')
        if acao == 'email':
            enviar_backup_email()
            if os.environ.get('SMTP_USER'):
                flash('E-mail de backup enviado!', 'success')
            else:
                flash('SMTP não configurado. Defina SMTP_USER e SMTP_PASS nas variáveis de ambiente.', 'warning')
            return redirect(url_for('admin_backup'))
        json_str = gerar_backup_json()
        nome = f"backup_ciclorh_{datetime.now().strftime('%Y%m%d_%H%M')}.json"
        _run("INSERT INTO config (chave, valor) VALUES ('ultimo_backup', %s) "
             "ON CONFLICT (chave) DO UPDATE SET valor = EXCLUDED.valor",
             (datetime.now().strftime('%d/%m/%Y às %H:%M'),))
        _log('backup', 'sistema', descricao='Download de backup realizado')
        return Response(
            json_str,
            mimetype='application/json',
            headers={'Content-Disposition': f'attachment; filename="{nome}"'}
        )
    ultimo = _q("SELECT valor FROM config WHERE chave='ultimo_backup'", one=True)
    smtp_ok = bool(os.environ.get('SMTP_USER') and os.environ.get('SMTP_PASS'))
    return render_template('admin/backup.html',
                           ultimo_backup=ultimo['valor'] if ultimo else None,
                           smtp_ok=smtp_ok,
                           backup_email=os.environ.get('BACKUP_EMAIL', os.environ.get('SMTP_USER', '—')))


@app.route('/admin/restaurar', methods=['POST'])
@admin_required
def admin_restaurar():
    arq = request.files.get('backup_file')
    if not arq or not arq.filename.endswith('.json'):
        flash('Selecione um arquivo .json exportado pelo sistema.', 'danger')
        return redirect(url_for('admin_backup'))
    try:
        conteudo = arq.read().decode('utf-8')
        restaurar_backup_json(conteudo)
        _log('restaurar', 'sistema', descricao='Banco restaurado via backup')
        flash('Banco de dados restaurado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro na restauração: {e}', 'danger')
    return redirect(url_for('admin_backup'))


# ─── ERROS ────────────────────────────────────────────────────────────────────

@app.errorhandler(403)
def forbidden(e):
    return render_template('erro.html', codigo=403,
                           msg='Acesso negado. Apenas o administrador pode acessar esta área.'), 403


@app.errorhandler(404)
def not_found(e):
    return render_template('erro.html', codigo=404, msg='Página não encontrada.'), 404


# ─── CRM ──────────────────────────────────────────────────────────────────────

def crm_required(f):
    @wraps(f)
    @login_required
    def decorated(*args, **kwargs):
        if not current_user.is_crm:
            abort(403)
        return f(*args, **kwargs)
    decorated.__name__ = f.__name__
    return decorated


def _crm_pode_ver_todos():
    return current_user.is_admin or current_user.crm_role == 'gerente'


def _crm_leads():
    if _crm_pode_ver_todos():
        return _q("""SELECT l.*, u.nome as resp_nome FROM crm_lead l
                     LEFT JOIN usuario u ON u.id = l.responsavel_id
                     ORDER BY l.updated_at DESC""")
    return _q("""SELECT l.*, u.nome as resp_nome FROM crm_lead l
                 LEFT JOIN usuario u ON u.id = l.responsavel_id
                 WHERE l.responsavel_id = %s
                 ORDER BY l.updated_at DESC""", (current_user.id,))


def _crm_usuarios():
    return _q("SELECT id, nome FROM usuario WHERE acesso_crm=TRUE OR role='admin' ORDER BY nome")


@app.route('/crm')
@crm_required
def crm_kanban():
    leads = _crm_leads()
    por_etapa = {e: [] for e in ETAPAS_CRM}
    for l in leads:
        if l['etapa'] in por_etapa:
            por_etapa[l['etapa']].append(l)
    return render_template('crm/kanban.html', por_etapa=por_etapa,
                           etapas=ETAPAS_CRM, cores=ETAPAS_CRM_COR,
                           total_leads=len(leads))


@app.route('/crm/lead/novo', methods=['GET', 'POST'])
@crm_required
def crm_lead_novo():
    if request.method == 'POST':
        resp_id = request.form.get('responsavel_id') or current_user.id
        _run("""INSERT INTO crm_lead
                (empresa_nome, empresa_cnpj, cidade, segmento, vagas_estimadas,
                 etapa, origem, responsavel_id, contato_nome, contato_email,
                 contato_whatsapp, obs)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (request.form['empresa_nome'].strip(),
              request.form.get('empresa_cnpj') or None,
              request.form.get('cidade') or None,
              request.form.get('segmento') or None,
              request.form.get('vagas_estimadas') or None,
              request.form.get('etapa', 'Lead Captado'),
              request.form.get('origem') or None,
              resp_id,
              request.form.get('contato_nome') or None,
              request.form.get('contato_email') or None,
              request.form.get('contato_whatsapp') or None,
              request.form.get('obs') or None))
        flash('Lead criado!', 'success')
        return redirect(url_for('crm_kanban'))
    return render_template('crm/form.html', lead=None, etapas=ETAPAS_CRM,
                           origens=ORIGENS_CRM, usuarios_crm=_crm_usuarios(),
                           pode_ver_todos=_crm_pode_ver_todos(),
                           current_user_id=int(current_user.id))


@app.route('/crm/lead/<int:id>')
@crm_required
def crm_lead_detalhe(id):
    lead = _q("""SELECT l.*, u.nome as resp_nome FROM crm_lead l
                 LEFT JOIN usuario u ON u.id = l.responsavel_id WHERE l.id=%s""", (id,), one=True)
    if not lead:
        abort(404)
    if not _crm_pode_ver_todos() and str(lead['responsavel_id']) != current_user.id:
        abort(403)
    interacoes = _q("""SELECT i.*, u.nome as autor FROM crm_interacao i
                       LEFT JOIN usuario u ON u.id = i.usuario_id
                       WHERE i.lead_id=%s ORDER BY i.created_at DESC""", (id,))
    etapa_idx = ETAPAS_CRM.index(lead['etapa']) if lead['etapa'] in ETAPAS_CRM else 0
    implantacao = _q("""SELECT i.*, COUNT(it.id) as total_itens,
                               COUNT(CASE WHEN it.concluido THEN 1 END) as itens_ok
                        FROM crm_implantacao i
                        LEFT JOIN crm_implantacao_item it ON it.implantacao_id = i.id
                        WHERE i.lead_id = %s
                        GROUP BY i.id LIMIT 1""", (id,), one=True)
    return render_template('crm/lead.html', lead=lead, interacoes=interacoes,
                           etapas=ETAPAS_CRM, etapa_idx=etapa_idx, cores=ETAPAS_CRM_COR,
                           tipos_interacao=TIPOS_INTERACAO, usuarios=_crm_usuarios(),
                           pode_ver_todos=_crm_pode_ver_todos(),
                           implantacao=implantacao)


@app.route('/crm/lead/<int:id>/editar', methods=['GET', 'POST'])
@crm_required
def crm_lead_editar(id):
    lead = _q("SELECT * FROM crm_lead WHERE id=%s", (id,), one=True)
    if not lead:
        abort(404)
    if not _crm_pode_ver_todos() and str(lead['responsavel_id']) != current_user.id:
        abort(403)
    if request.method == 'POST':
        resp_id = request.form.get('responsavel_id') or lead['responsavel_id']
        _run("""UPDATE crm_lead SET empresa_nome=%s, empresa_cnpj=%s, cidade=%s, segmento=%s,
                vagas_estimadas=%s, etapa=%s, origem=%s, responsavel_id=%s,
                contato_nome=%s, contato_email=%s, contato_whatsapp=%s, obs=%s,
                updated_at=NOW() WHERE id=%s""",
             (request.form['empresa_nome'].strip(),
              request.form.get('empresa_cnpj') or None,
              request.form.get('cidade') or None,
              request.form.get('segmento') or None,
              request.form.get('vagas_estimadas') or None,
              request.form.get('etapa', lead['etapa']),
              request.form.get('origem') or None,
              resp_id,
              request.form.get('contato_nome') or None,
              request.form.get('contato_email') or None,
              request.form.get('contato_whatsapp') or None,
              request.form.get('obs') or None,
              id))
        flash('Lead atualizado!', 'success')
        return redirect(url_for('crm_lead_detalhe', id=id))
    return render_template('crm/form.html', lead=lead, etapas=ETAPAS_CRM,
                           origens=ORIGENS_CRM, usuarios_crm=_crm_usuarios(),
                           pode_ver_todos=_crm_pode_ver_todos(),
                           current_user_id=int(current_user.id))


@app.route('/crm/lead/<int:id>/etapa', methods=['POST'])
@crm_required
def crm_lead_etapa(id):
    nova = request.form.get('etapa')
    if nova not in ETAPAS_CRM:
        abort(400)
    _run("UPDATE crm_lead SET etapa=%s, updated_at=NOW() WHERE id=%s", (nova, id))
    flash(f'Etapa alterada para <strong>{nova}</strong>.', 'success')
    if nova == 'Cliente Ativo':
        lead = _q("SELECT * FROM crm_lead WHERE id=%s", (id,), one=True)
        resp_id = lead['responsavel_id'] if lead else current_user.id
        imp_id = _criar_implantacao(id, resp_id)
        flash(f'Pipeline de implantação criado. <a href="{url_for("crm_implantacao_detalhe", id=imp_id)}" class="alert-link">Ver pipeline →</a>', 'info')
    return redirect(url_for('crm_lead_detalhe', id=id))


@app.route('/crm/lead/<int:id>/interacao', methods=['POST'])
@crm_required
def crm_interacao_nova(id):
    tipo = request.form.get('tipo', '').strip()
    descricao = request.form.get('descricao', '').strip()
    if not tipo or not descricao:
        flash('Tipo e descrição são obrigatórios.', 'danger')
        return redirect(url_for('crm_lead_detalhe', id=id))
    _run("INSERT INTO crm_interacao (lead_id, tipo, descricao, usuario_id) VALUES (%s,%s,%s,%s)",
         (id, tipo, descricao, current_user.id))
    _run("UPDATE crm_lead SET updated_at=NOW() WHERE id=%s", (id,))
    flash('Interação registrada.', 'success')
    return redirect(url_for('crm_lead_detalhe', id=id))


@app.route('/crm/lead/<int:id>/excluir', methods=['POST'])
@crm_required
def crm_lead_excluir(id):
    if not _crm_pode_ver_todos():
        abort(403)
    _run("DELETE FROM crm_lead WHERE id=%s", (id,))
    flash('Lead excluído.', 'warning')
    return redirect(url_for('crm_kanban'))


@app.route('/crm/calculadora')
@crm_required
def crm_calculadora():
    return render_template('crm/calculadora_custo.html')


@app.route('/crm/indicadores')
@crm_required
def crm_indicadores():
    leads = _crm_leads()
    total_leads = len(leads)
    por_etapa_count = {e: 0 for e in ETAPAS_CRM}
    for l in leads:
        if l['etapa'] in por_etapa_count:
            por_etapa_count[l['etapa']] += 1
    clientes_ativos = por_etapa_count.get('Cliente Ativo', 0)
    contratos_enviados = por_etapa_count.get('Contrato Enviado', 0) + por_etapa_count.get('Contrato Assinado', 0)
    taxa_conversao = round(clientes_ativos / total_leads * 100, 1) if total_leads else 0

    pode_ver = _crm_pode_ver_todos()
    por_origem = _q("""SELECT origem, COUNT(*) as total FROM crm_lead
                       GROUP BY origem ORDER BY total DESC""")
    por_responsavel = None
    if pode_ver:
        por_responsavel = _q("""SELECT u.nome as resp_nome,
                                COUNT(l.id) as total,
                                COUNT(CASE WHEN l.etapa='Cliente Ativo' THEN 1 END) as clientes
                                FROM crm_lead l
                                LEFT JOIN usuario u ON u.id = l.responsavel_id
                                GROUP BY u.nome ORDER BY total DESC""")
    ultimas_interacoes = _q("""SELECT i.*, l.empresa_nome, l.id as lead_id
                               FROM crm_interacao i
                               JOIN crm_lead l ON l.id = i.lead_id
                               ORDER BY i.created_at DESC LIMIT 10""")
    return render_template('crm/indicadores.html',
                           total_leads=total_leads, por_etapa_count=por_etapa_count,
                           etapas=ETAPAS_CRM, cores=ETAPAS_CRM_COR,
                           clientes_ativos=clientes_ativos,
                           contratos_enviados=contratos_enviados,
                           taxa_conversao=taxa_conversao,
                           por_origem=por_origem, por_responsavel=por_responsavel,
                           pode_ver_todos=pode_ver,
                           ultimas_interacoes=ultimas_interacoes)



# ─── CRM — PIPELINE DE IMPLANTAÇÃO ────────────────────────────────────────────

def _criar_implantacao(lead_id, responsavel_id):
    """Cria pipeline de implantação com checklist padrão. Ignora se já existir."""
    existente = _q("SELECT id FROM crm_implantacao WHERE lead_id=%s", (lead_id,), one=True)
    if existente:
        return existente['id']
    _run("INSERT INTO crm_implantacao (lead_id, responsavel_id) VALUES (%s, %s)",
         (lead_id, responsavel_id))
    imp = _q("SELECT id FROM crm_implantacao WHERE lead_id=%s ORDER BY id DESC LIMIT 1",
             (lead_id,), one=True)
    imp_id = imp['id']
    for i, titulo in enumerate(ITENS_IMPLANTACAO_PADRAO):
        _run("INSERT INTO crm_implantacao_item (implantacao_id, titulo, ordem) VALUES (%s,%s,%s)",
             (imp_id, titulo, i))
    return imp_id


@app.route('/crm/implantacao')
@crm_required
def crm_implantacao_lista():
    if _crm_pode_ver_todos():
        implantacoes = _q("""
            SELECT i.*, l.empresa_nome, l.cidade, u.nome as resp_nome,
                   COUNT(it.id) as total_itens,
                   COUNT(CASE WHEN it.concluido THEN 1 END) as itens_ok
            FROM crm_implantacao i
            JOIN crm_lead l ON l.id = i.lead_id
            LEFT JOIN usuario u ON u.id = i.responsavel_id
            LEFT JOIN crm_implantacao_item it ON it.implantacao_id = i.id
            GROUP BY i.id, l.empresa_nome, l.cidade, u.nome
            ORDER BY i.updated_at DESC""")
    else:
        implantacoes = _q("""
            SELECT i.*, l.empresa_nome, l.cidade, u.nome as resp_nome,
                   COUNT(it.id) as total_itens,
                   COUNT(CASE WHEN it.concluido THEN 1 END) as itens_ok
            FROM crm_implantacao i
            JOIN crm_lead l ON l.id = i.lead_id
            LEFT JOIN usuario u ON u.id = i.responsavel_id
            LEFT JOIN crm_implantacao_item it ON it.implantacao_id = i.id
            WHERE i.responsavel_id = %s
            GROUP BY i.id, l.empresa_nome, l.cidade, u.nome
            ORDER BY i.updated_at DESC""", (current_user.id,))
    return render_template('crm/implantacao_lista.html', implantacoes=implantacoes,
                           pode_ver_todos=_crm_pode_ver_todos())


@app.route('/crm/implantacao/<int:id>')
@crm_required
def crm_implantacao_detalhe(id):
    imp = _q("""SELECT i.*, l.empresa_nome, l.cidade, l.id as lead_id,
                       u.nome as resp_nome
                FROM crm_implantacao i
                JOIN crm_lead l ON l.id = i.lead_id
                LEFT JOIN usuario u ON u.id = i.responsavel_id
                WHERE i.id = %s""", (id,), one=True)
    if not imp:
        abort(404)
    if not _crm_pode_ver_todos() and str(imp['responsavel_id']) != current_user.id:
        abort(403)
    itens = _q("""SELECT it.*, u.nome as concluido_por_nome
                  FROM crm_implantacao_item it
                  LEFT JOIN usuario u ON u.id = it.concluido_por
                  WHERE it.implantacao_id = %s ORDER BY it.ordem, it.id""", (id,))
    total = len(itens)
    concluidos = sum(1 for it in itens if it['concluido'])
    pct = round(concluidos / total * 100) if total else 0
    return render_template('crm/implantacao_detalhe.html', imp=imp, itens=itens,
                           total=total, concluidos=concluidos, pct=pct,
                           pode_ver_todos=_crm_pode_ver_todos(),
                           usuarios_crm=_crm_usuarios())


@app.route('/crm/implantacao/<int:id>/item/<int:item_id>/toggle', methods=['POST'])
@crm_required
def crm_implantacao_item_toggle(id, item_id):
    item = _q("SELECT * FROM crm_implantacao_item WHERE id=%s AND implantacao_id=%s",
              (item_id, id), one=True)
    if not item:
        abort(404)
    novo = not item['concluido']
    if novo:
        _run("""UPDATE crm_implantacao_item SET concluido=TRUE,
                concluido_em=NOW(), concluido_por=%s WHERE id=%s""",
             (current_user.id, item_id))
    else:
        _run("""UPDATE crm_implantacao_item SET concluido=FALSE,
                concluido_em=NULL, concluido_por=NULL WHERE id=%s""", (item_id,))
    _run("UPDATE crm_implantacao SET updated_at=NOW() WHERE id=%s", (id,))
    return redirect(url_for('crm_implantacao_detalhe', id=id))


@app.route('/crm/implantacao/<int:id>/fechar', methods=['POST'])
@crm_required
def crm_implantacao_fechar(id):
    _run("UPDATE crm_implantacao SET status='concluida', updated_at=NOW() WHERE id=%s", (id,))
    flash('Implantação marcada como concluída!', 'success')
    return redirect(url_for('crm_implantacao_lista'))


@app.route('/crm/implantacao/<int:id>/responsavel', methods=['POST'])
@crm_required
def crm_implantacao_responsavel(id):
    if not _crm_pode_ver_todos():
        abort(403)
    resp_id = request.form.get('responsavel_id')
    _run("UPDATE crm_implantacao SET responsavel_id=%s, updated_at=NOW() WHERE id=%s",
         (resp_id or None, id))
    flash('Responsável atualizado.', 'success')
    return redirect(url_for('crm_implantacao_detalhe', id=id))


@app.route('/crm/implantacao/<int:id>/obs', methods=['POST'])
@crm_required
def crm_implantacao_obs(id):
    _run("UPDATE crm_implantacao SET obs=%s, updated_at=NOW() WHERE id=%s",
         (request.form.get('obs') or None, id))
    flash('Observação salva.', 'success')
    return redirect(url_for('crm_implantacao_detalhe', id=id))


# ─── GESTÃO DE VAGAS ──────────────────────────────────────────────────────────

STATUS_VAGA = ['aberta', 'em_selecao', 'preenchida', 'cancelada']
STATUS_VAGA_COR = {'aberta': 'success', 'em_selecao': 'primary',
                   'preenchida': 'secondary', 'cancelada': 'danger'}
STATUS_CANDIDATURA = ['inscrito', 'em_entrevista', 'aprovado', 'reprovado',
                      'nao_compareceu_j', 'nao_compareceu_nj', 'desistiu']
STATUS_CANDIDATURA_COR = {'inscrito': 'secondary', 'em_entrevista': 'primary',
                           'aprovado': 'success', 'reprovado': 'danger',
                           'nao_compareceu_j': 'warning', 'nao_compareceu_nj': 'danger',
                           'desistiu': 'warning'}


# ── Vagas ──────────────────────────────────────────────────────────────────────

@app.route('/vagas')
@login_required
def vagas_lista():
    status_f = request.args.get('status', '')
    empresa_f = request.args.get('empresa_id', '')
    q = """SELECT v.*, e.nome as emp_nome, a.nome as area_nome,
                  u.nome as resp_nome,
                  (SELECT COUNT(*) FROM candidatura c WHERE c.vaga_id = v.id) as total_cands,
                  (SELECT COUNT(*) FROM candidatura c WHERE c.vaga_id = v.id AND c.status='aprovado') as aprovados
           FROM vaga v
           LEFT JOIN empresa e ON e.id = v.empresa_id
           LEFT JOIN area_estagio a ON a.id = v.area_id
           LEFT JOIN usuario u ON u.id = v.responsavel_id
           WHERE 1=1"""
    params = []
    if status_f:
        q += " AND v.status = %s"; params.append(status_f)
    if empresa_f:
        q += " AND v.empresa_id = %s"; params.append(empresa_f)
    q += " ORDER BY v.created_at DESC"
    vagas = _q(q, params or None)
    empresas = _q("SELECT id, nome, COALESCE(nome_fantasia, nome) AS display FROM empresa ORDER BY COALESCE(nome_fantasia, nome)")
    return render_template('vagas/lista.html', vagas=vagas, empresas=empresas,
                           status_f=status_f, empresa_f=empresa_f,
                           status_cor=STATUS_VAGA_COR)


@app.route('/vagas/nova', methods=['GET', 'POST'])
@login_required
def vaga_nova():
    if request.method == 'POST':
        _run("""INSERT INTO vaga (empresa_id, area_id, titulo, descricao, requisitos,
                curso_desejado, nivel, carga_horaria, bolsa, beneficios,
                vagas_total, data_limite, responsavel_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (request.form.get('empresa_id') or None,
              request.form.get('area_id') or None,
              request.form['titulo'].strip(),
              request.form.get('descricao') or None,
              request.form.get('requisitos') or None,
              request.form.get('curso_desejado') or None,
              request.form.get('nivel', 'superior'),
              request.form.get('carga_horaria') or None,
              request.form.get('bolsa') or None,
              request.form.get('beneficios') or None,
              request.form.get('vagas_total') or 1,
              request.form.get('data_limite') or None,
              current_user.id))
        flash('Vaga criada!', 'success')
        return redirect(url_for('vagas_lista'))
    empresas = _q("SELECT id, nome, COALESCE(nome_fantasia, nome) AS display FROM empresa ORDER BY COALESCE(nome_fantasia, nome)")
    areas = _q("SELECT id, nome FROM area_estagio ORDER BY nome")
    return render_template('vagas/form.html', vaga=None, empresas=empresas, areas=areas)


@app.route('/vagas/<int:id>')
@login_required
def vaga_detalhe(id):
    vaga = _q("""SELECT v.*, e.nome as emp_nome, e.bairro as emp_bairro, e.cidade as emp_cidade,
                        a.nome as area_nome, u.nome as resp_nome
                 FROM vaga v
                 LEFT JOIN empresa e ON e.id = v.empresa_id
                 LEFT JOIN area_estagio a ON a.id = v.area_id
                 LEFT JOIN usuario u ON u.id = v.responsavel_id
                 WHERE v.id = %s""", (id,), one=True)
    if not vaga:
        abort(404)
    candidaturas = _q("""SELECT c.*, ca.nome as cand_nome, ca.curso, ca.whatsapp,
                                ca.disponibilidade, ca.id as cand_id
                          FROM candidatura c
                          JOIN candidato ca ON ca.id = c.candidato_id
                          WHERE c.vaga_id = %s ORDER BY c.created_at""", (id,))
    # candidatos disponíveis: sem candidatura ativa em QUALQUER vaga
    # e sem candidatura ativa (inscrito/em_entrevista/aprovado) NESTA vaga
    # (permite re-encaminhar candidatos com status terminal nesta vaga)
    disponiveis = _q("""SELECT id, nome, curso, whatsapp FROM candidato
                        WHERE id NOT IN (
                            SELECT candidato_id FROM candidatura
                            WHERE status IN ('inscrito','em_entrevista')
                        )
                        AND id NOT IN (
                            SELECT candidato_id FROM candidatura
                            WHERE vaga_id = %s
                            AND status IN ('inscrito','em_entrevista','aprovado')
                        )
                        ORDER BY nome""", (id,))
    return render_template('vagas/detalhe.html', vaga=vaga, candidaturas=candidaturas,
                           disponiveis=disponiveis,
                           status_cor=STATUS_CANDIDATURA_COR,
                           status_vaga_cor=STATUS_VAGA_COR)


@app.route('/candidato/cadastro', methods=['GET', 'POST'])
def candidato_cadastro_publico():
    ies = _q("SELECT id, nome FROM ie ORDER BY nome")
    if request.method == 'POST':
        nome = request.form.get('nome', '').strip()
        if not nome:
            return render_template('candidatos/cadastro_publico.html', erro='Nome é obrigatório.', ies=ies)

        email       = request.form.get('email', '').strip() or None
        whatsapp    = request.form.get('whatsapp', '').strip() or None
        cpf         = request.form.get('cpf', '').strip() or None
        dt_nasc     = request.form.get('data_nascimento', '').strip() or None
        endereco    = request.form.get('endereco', '').strip() or None
        bairro      = request.form.get('bairro', '').strip() or None
        cidade      = request.form.get('cidade', '').strip() or None
        estado      = request.form.get('estado', '').strip() or None
        curso       = request.form.get('curso', '').strip() or None
        semestre    = request.form.get('semestre', '').strip() or None
        ie_id       = request.form.get('ie_id') or None
        disponib    = request.form.get('disponibilidade') or None
        obs         = request.form.get('obs', '').strip() or None

        cand = None
        if email:
            cand = _q("SELECT id FROM candidato WHERE email=%s", (email,), one=True)
        if not cand and whatsapp:
            cand = _q("SELECT id FROM candidato WHERE whatsapp=%s", (whatsapp,), one=True)

        if not cand:
            cand_id = _ins("""INSERT INTO candidato
                              (nome, cpf, email, whatsapp, data_nascimento,
                               endereco, bairro, cidade, estado,
                               curso, semestre, ie_id, disponibilidade, obs)
                              VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
                           (nome, cpf, email, whatsapp, dt_nasc,
                            endereco, bairro, cidade, estado,
                            curso, semestre, ie_id, disponib, obs))
            if request.form.get('ja_trabalhou') == 'sim':
                _save_experiencias(cand_id)

        return render_template('candidatos/cadastro_sucesso.html', nome=nome)

    return render_template('candidatos/cadastro_publico.html', erro=None, ies=ies)


@app.route('/api/candidato/<int:cand_id>/historico-empresa/<int:empresa_id>')
@login_required
def api_candidato_historico_empresa(cand_id, empresa_id):
    hist = _q("""SELECT ca.status, v.titulo, ca.created_at
                 FROM candidatura ca
                 JOIN vaga v ON v.id = ca.vaga_id
                 WHERE ca.candidato_id = %s AND v.empresa_id = %s
                 AND ca.status = 'reprovado'
                 ORDER BY ca.created_at DESC""", (cand_id, empresa_id))
    return jsonify({'historico': [{'titulo': r['titulo'],
                                   'data': fmt_date(r['created_at'])} for r in hist]})


@app.route('/candidatos/<int:id>/curriculo')
@login_required
def candidato_curriculo(id):
    c = _q("""SELECT ca.*, i.nome as ie_nome
              FROM candidato ca
              LEFT JOIN ie i ON i.id = ca.ie_id
              WHERE ca.id = %s""", (id,), one=True)
    if not c:
        abort(404)
    experiencias = _q("SELECT * FROM candidato_experiencia WHERE candidato_id=%s ORDER BY ordem", (id,))
    return render_template('candidatos/curriculo.html', candidato=c, experiencias=experiencias)


@app.route('/vagas/<int:id>/curriculos')
@login_required
def vaga_curriculos(id):
    vaga = _q("""SELECT v.*, e.nome as emp_nome
                 FROM vaga v LEFT JOIN empresa e ON e.id = v.empresa_id
                 WHERE v.id=%s""", (id,), one=True)
    if not vaga:
        abort(404)
    candidatos_raw = _q("""
        SELECT ca.*, i.nome as ie_nome, cu.status as status_candidatura, cu.id as cand_id
        FROM candidatura cu
        JOIN candidato ca ON ca.id = cu.candidato_id
        LEFT JOIN ie i ON i.id = ca.ie_id
        WHERE cu.vaga_id = %s AND cu.status IN ('inscrito','em_entrevista','aprovado')
        ORDER BY ca.nome""", (id,))
    candidatos = []
    for c in candidatos_raw:
        exp = _q("SELECT * FROM candidato_experiencia WHERE candidato_id=%s ORDER BY ordem", (c['id'],))
        candidatos.append({'candidato': c, 'experiencias': exp})
    return render_template('vagas/curriculos.html', vaga=vaga, candidatos=candidatos)


@app.route('/vagas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def vaga_editar(id):
    vaga = _q("SELECT * FROM vaga WHERE id=%s", (id,), one=True)
    if not vaga:
        abort(404)
    if request.method == 'POST':
        _run("""UPDATE vaga SET empresa_id=%s, area_id=%s, titulo=%s, descricao=%s,
                requisitos=%s, curso_desejado=%s, nivel=%s, carga_horaria=%s,
                bolsa=%s, beneficios=%s, vagas_total=%s, data_limite=%s,
                responsavel_id=%s, updated_at=NOW() WHERE id=%s""",
             (request.form.get('empresa_id') or None,
              request.form.get('area_id') or None,
              request.form['titulo'].strip(),
              request.form.get('descricao') or None,
              request.form.get('requisitos') or None,
              request.form.get('curso_desejado') or None,
              request.form.get('nivel', 'superior'),
              request.form.get('carga_horaria') or None,
              request.form.get('bolsa') or None,
              request.form.get('beneficios') or None,
              request.form.get('vagas_total') or 1,
              request.form.get('data_limite') or None,
              request.form.get('responsavel_id') or None,
              id))
        flash('Vaga atualizada!', 'success')
        return redirect(url_for('vaga_detalhe', id=id))
    empresas = _q("SELECT id, nome, COALESCE(nome_fantasia, nome) AS display FROM empresa ORDER BY COALESCE(nome_fantasia, nome)")
    areas = _q("SELECT id, nome FROM area_estagio ORDER BY nome")
    usuarios = _q("SELECT id, nome FROM usuario ORDER BY nome")
    return render_template('vagas/form.html', vaga=vaga, empresas=empresas,
                           areas=areas, usuarios=usuarios)


@app.route('/vagas/<int:id>/excluir', methods=['POST'])
@login_required
def vaga_excluir(id):
    v = _q("SELECT titulo FROM vaga WHERE id=%s", (id,), one=True)
    if not v:
        abort(404)
    _run("DELETE FROM vaga WHERE id=%s", (id,))
    _log('excluir', 'vaga', id, f'Excluiu vaga: {v["titulo"]}')
    flash(f'Vaga "{v["titulo"]}" excluída.', 'warning')
    return redirect(url_for('vagas_lista'))


@app.route('/vagas/<int:id>/status', methods=['POST'])
@login_required
def vaga_status(id):
    novo = request.form.get('status')
    if novo not in STATUS_VAGA:
        abort(400)
    _run("UPDATE vaga SET status=%s, updated_at=NOW() WHERE id=%s", (novo, id))
    flash(f'Status da vaga alterado para <strong>{novo}</strong>.', 'success')
    return redirect(url_for('vaga_detalhe', id=id))


# ── Candidatos ─────────────────────────────────────────────────────────────────

@app.route('/candidatos')
@login_required
def candidatos_lista():
    busca = request.args.get('q', '').strip()
    q = """SELECT c.*, i.nome as ie_nome,
                  (SELECT COUNT(*) FROM candidatura ca WHERE ca.candidato_id = c.id) as num_candidaturas,
                  (SELECT COUNT(*) FROM candidatura ca WHERE ca.candidato_id = c.id AND ca.status='aprovado') as aprovados
           FROM candidato c LEFT JOIN ie i ON i.id = c.ie_id WHERE 1=1"""
    params = []
    if busca:
        q += " AND (c.nome ILIKE %s OR c.curso ILIKE %s OR c.cidade ILIKE %s OR i.nome ILIKE %s)"
        params += [f'%{busca}%', f'%{busca}%', f'%{busca}%', f'%{busca}%']
    q += " ORDER BY c.created_at DESC"
    candidatos = _q(q, params or None)
    return render_template('candidatos/lista.html', candidatos=candidatos, busca=busca)


def _save_experiencias(candidato_id):
    _run("DELETE FROM candidato_experiencia WHERE candidato_id=%s", (candidato_id,))
    empresas  = request.form.getlist('exp_empresa[]')
    periodos  = request.form.getlist('exp_periodo[]')
    funcoes   = request.form.getlist('exp_funcao[]')
    for i, emp in enumerate(empresas):
        if emp.strip():
            _run("""INSERT INTO candidato_experiencia (candidato_id, empresa, periodo, funcao, ordem)
                    VALUES (%s,%s,%s,%s,%s)""",
                 (candidato_id, emp.strip(),
                  periodos[i].strip() if i < len(periodos) else '',
                  funcoes[i].strip()  if i < len(funcoes)  else '',
                  i))


@app.route('/candidatos/novo', methods=['GET', 'POST'])
@login_required
def candidato_novo():
    if request.method == 'POST':
        cand_id = _ins("""INSERT INTO candidato (nome, cpf, email, whatsapp, data_nascimento,
                endereco, bairro, cidade, estado, curso, semestre, ie_id, disponibilidade, obs)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (request.form['nome'].strip(),
              request.form.get('cpf') or None,
              request.form.get('email') or None,
              request.form.get('whatsapp') or None,
              request.form.get('data_nascimento') or None,
              request.form.get('endereco') or None,
              request.form.get('bairro') or None,
              request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('curso') or None,
              request.form.get('semestre') or None,
              request.form.get('ie_id') or None,
              request.form.get('disponibilidade') or None,
              request.form.get('obs') or None))
        if request.form.get('ja_trabalhou') == 'sim':
            _save_experiencias(cand_id)
        flash('Candidato cadastrado!', 'success')
        return redirect(url_for('candidatos_lista'))
    ies = _q("SELECT id, nome FROM ie ORDER BY nome")
    return render_template('candidatos/form.html', candidato=None, experiencias=[], ies=ies)


@app.route('/candidatos/<int:id>')
@login_required
def candidato_detalhe(id):
    c = _q("""SELECT ca.*, i.nome as ie_nome FROM candidato ca
              LEFT JOIN ie i ON i.id = ca.ie_id WHERE ca.id = %s""", (id,), one=True)
    if not c:
        abort(404)
    candidaturas = _q("""SELECT cu.*, v.titulo as vaga_titulo, e.nome as emp_nome, cu.id as cand_id
                          FROM candidatura cu
                          JOIN vaga v ON v.id = cu.vaga_id
                          LEFT JOIN empresa e ON e.id = v.empresa_id
                          WHERE cu.candidato_id = %s ORDER BY cu.created_at DESC""", (id,))
    vagas_disponiveis = _q("""SELECT v.id, v.titulo, e.nome as emp_nome
                               FROM vaga v LEFT JOIN empresa e ON e.id = v.empresa_id
                               WHERE v.status IN ('aberta','em_selecao')
                               AND v.id NOT IN (
                                   SELECT vaga_id FROM candidatura WHERE candidato_id = %s
                               ) ORDER BY v.titulo""", (id,))
    experiencias = _q("SELECT * FROM candidato_experiencia WHERE candidato_id=%s ORDER BY ordem", (id,))
    return render_template('candidatos/detalhe.html', candidato=c, candidaturas=candidaturas,
                           vagas_disponiveis=vagas_disponiveis,
                           experiencias=experiencias,
                           status_cor=STATUS_CANDIDATURA_COR)


@app.route('/candidatos/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def candidato_editar(id):
    c = _q("SELECT * FROM candidato WHERE id=%s", (id,), one=True)
    if not c:
        abort(404)
    if request.method == 'POST':
        _run("""UPDATE candidato SET nome=%s, cpf=%s, email=%s, whatsapp=%s,
                data_nascimento=%s, endereco=%s, bairro=%s, cidade=%s, estado=%s,
                curso=%s, semestre=%s, ie_id=%s, disponibilidade=%s, obs=%s WHERE id=%s""",
             (request.form['nome'].strip(),
              request.form.get('cpf') or None,
              request.form.get('email') or None,
              request.form.get('whatsapp') or None,
              request.form.get('data_nascimento') or None,
              request.form.get('endereco') or None,
              request.form.get('bairro') or None,
              request.form.get('cidade') or None,
              request.form.get('estado') or None,
              request.form.get('curso') or None,
              request.form.get('semestre') or None,
              request.form.get('ie_id') or None,
              request.form.get('disponibilidade') or None,
              request.form.get('obs') or None,
              id))
        if request.form.get('ja_trabalhou') == 'sim':
            _save_experiencias(id)
        else:
            _run("DELETE FROM candidato_experiencia WHERE candidato_id=%s", (id,))
        flash('Candidato atualizado!', 'success')
        return redirect(url_for('candidato_detalhe', id=id))
    ies = _q("SELECT id, nome FROM ie ORDER BY nome")
    experiencias = _q("SELECT * FROM candidato_experiencia WHERE candidato_id=%s ORDER BY ordem", (id,))
    return render_template('candidatos/form.html', candidato=c, experiencias=experiencias, ies=ies)


@app.route('/candidatos/<int:id>/excluir', methods=['POST'])
@login_required
def candidato_excluir(id):
    c = _q("SELECT nome FROM candidato WHERE id=%s", (id,), one=True)
    if not c:
        abort(404)
    _run("DELETE FROM candidato WHERE id=%s", (id,))
    _log('excluir', 'candidato', id, f'Excluiu candidato: {c["nome"]}')
    flash(f'Candidato "{c["nome"]}" excluído.', 'warning')
    return redirect(url_for('candidatos_lista'))


# ── Candidaturas ───────────────────────────────────────────────────────────────

@app.route('/candidatura/nova', methods=['POST'])
@login_required
def candidatura_nova():
    vaga_id = request.form.get('vaga_id')
    candidato_id = request.form.get('candidato_id')
    origem = request.form.get('origem', 'vaga')

    existente = _q("""SELECT id, status FROM candidatura
                      WHERE vaga_id=%s AND candidato_id=%s""",
                   (vaga_id, candidato_id), one=True)

    if existente:
        if existente['status'] in ('inscrito', 'em_entrevista', 'aprovado'):
            flash('Este candidato já está ativo nesta vaga.', 'warning')
        else:
            # Nova rodada: reativa a candidatura existente
            _run("""UPDATE candidatura SET status='inscrito', obs=NULL, updated_at=NOW()
                    WHERE id=%s""", (existente['id'],))
            flash('Candidato reencaminhado para nova rodada!', 'success')
    else:
        _run("INSERT INTO candidatura (vaga_id, candidato_id) VALUES (%s,%s)",
             (vaga_id, candidato_id))
        flash('Candidato inscrito na vaga!', 'success')

    if origem == 'candidato':
        return redirect(url_for('candidato_detalhe', id=candidato_id))
    return redirect(url_for('vaga_detalhe', id=vaga_id))


@app.route('/candidatura/<int:id>/status', methods=['POST'])
@login_required
def candidatura_status(id):
    novo = request.form.get('status')
    if novo not in STATUS_CANDIDATURA:
        abort(400)
    c = _q("SELECT * FROM candidatura WHERE id=%s", (id,), one=True)
    if not c:
        abort(404)

    # Regra: 2ª falta sem justificativa → exclusão automática do candidato
    if novo == 'nao_compareceu_nj':
        prev = _q("""SELECT COUNT(*) as n FROM candidatura
                     WHERE candidato_id=%s AND status='nao_compareceu_nj' AND id != %s""",
                  (c['candidato_id'], id), one=True)
        _run("UPDATE candidatura SET status=%s, updated_at=NOW() WHERE id=%s", (novo, id))
        _run("UPDATE vaga SET updated_at=NOW() WHERE id=%s", (c['vaga_id'],))
        if prev and prev['n'] >= 1:
            info = _q("SELECT nome FROM candidato WHERE id=%s", (c['candidato_id'],), one=True)
            nome_cand = info['nome'] if info else f'Candidato {c["candidato_id"]}'
            _run("DELETE FROM candidato WHERE id=%s", (c['candidato_id'],))
            _log('excluir', 'candidato', c['candidato_id'],
                 f'Auto-excluído por 2 faltas sem justificativa: {nome_cand}')
            flash(f'<strong>{nome_cand}</strong> foi excluído do banco por 2 faltas sem justificativa.', 'danger')
        else:
            flash('Falta registrada. Na próxima falta sem justificativa o candidato será excluído automaticamente.', 'warning')
        return redirect(url_for('vaga_detalhe', id=c['vaga_id']))

    _run("UPDATE candidatura SET status=%s, updated_at=NOW() WHERE id=%s", (novo, id))
    _run("UPDATE vaga SET updated_at=NOW() WHERE id=%s", (c['vaga_id'],))
    if novo == 'aprovado':
        vaga = _q("SELECT empresa_id FROM vaga WHERE id=%s", (c['vaga_id'],), one=True)
        link = url_for('contrato_novo')
        if vaga and vaga['empresa_id']:
            link += f'?empresa_id={vaga["empresa_id"]}'
        flash(f'Candidato aprovado! <a href="{link}" class="alert-link">Criar contrato →</a>', 'success')
    else:
        flash('Status atualizado.', 'info')
    return redirect(url_for('vaga_detalhe', id=c['vaga_id']))


@app.route('/candidatura/<int:id>/obs', methods=['POST'])
@login_required
def candidatura_obs(id):
    c = _q("SELECT * FROM candidatura WHERE id=%s", (id,), one=True)
    if not c:
        abort(404)
    _run("UPDATE candidatura SET obs=%s, updated_at=NOW() WHERE id=%s",
         (request.form.get('obs') or None, id))
    return redirect(url_for('vaga_detalhe', id=c['vaga_id']))


# ─── Relatórios Gerenciais ────────────────────────────────────────────────────

def _xlsx_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.read(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment; filename="{filename}"'})


def _xlsx_header_style():
    fill = PatternFill('solid', start_color='1A3E6C', end_color='1A3E6C')
    font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
    align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(style='thin', color='DDDDDD')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    return fill, font, align, border


def _xlsx_apply_header(ws, headers):
    fill, font, align, border = _xlsx_header_style()
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        cell.border = border
    ws.row_dimensions[1].height = 28


def _xlsx_fmt_date(v):
    if v is None:
        return ''
    if hasattr(v, 'strftime'):
        return v.strftime('%d/%m/%Y')
    return str(v)


@app.route('/manual')
@login_required
def manual():
    return render_template('manual.html')


@app.route('/relatorios')
@login_required
def relatorios():
    return render_template('relatorios/index.html')


@app.route('/relatorios/contratos.xlsx')
@login_required
def relatorio_contratos_xlsx():
    situacao = request.args.get('situacao', 'ativos')
    if situacao == 'ativos':
        rows = _q("""
            SELECT c.id, e.nome as estagiario, emp.nome as empresa, ie.nome as ie,
                   c.data_inicio, c.data_termino, c.carga_horaria, c.valor_bolsa,
                   c.status, c.data_encerramento, a.nome as area
            FROM contrato c
            JOIN estagiario e ON e.id=c.estagiario_id
            JOIN empresa emp ON emp.id=c.empresa_id
            JOIN ie ON ie.id=c.ie_id
            LEFT JOIN area_estagio a ON a.id=c.area_id
            WHERE c.data_encerramento IS NULL
            ORDER BY e.nome""")
    else:
        rows = _q("""
            SELECT c.id, e.nome as estagiario, emp.nome as empresa, ie.nome as ie,
                   c.data_inicio, c.data_termino, c.carga_horaria, c.valor_bolsa,
                   c.status, c.data_encerramento, a.nome as area
            FROM contrato c
            JOIN estagiario e ON e.id=c.estagiario_id
            JOIN empresa emp ON emp.id=c.empresa_id
            JOIN ie ON ie.id=c.ie_id
            LEFT JOIN area_estagio a ON a.id=c.area_id
            ORDER BY c.data_inicio DESC""")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Contratos'
    headers = ['ID', 'Estagiário', 'Empresa', 'IE', 'Área', 'Início', 'Término',
               'CH (h/sem)', 'Bolsa (R$)', 'Status', 'Encerrado em']
    _xlsx_apply_header(ws, headers)

    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill('solid', start_color='F2F6FC', end_color='F2F6FC')
    for i, r in enumerate(rows, 2):
        vals = [r['id'], r['estagiario'], r['empresa'], r['ie'], r['area'] or '',
                _xlsx_fmt_date(r['data_inicio']), _xlsx_fmt_date(r['data_termino']),
                r['carga_horaria'], float(r['valor_bolsa']) if r['valor_bolsa'] else '',
                r['status'] or '', _xlsx_fmt_date(r['data_encerramento'])]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = data_font
            if i % 2 == 0:
                cell.fill = alt_fill

    col_widths = [6, 30, 30, 25, 20, 13, 13, 10, 12, 14, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'

    label = 'ativos' if situacao == 'ativos' else 'todos'
    return _xlsx_response(wb, f'contratos_{label}_{date.today()}.xlsx')


@app.route('/relatorios/estagiarios.xlsx')
@login_required
def relatorio_estagiarios_xlsx():
    rows = _q("""
        SELECT e.id, e.nome, e.cpf, e.email, e.telefone,
               e.cidade, e.estado, e.curso, e.semestre,
               (SELECT COUNT(*) FROM contrato WHERE estagiario_id=e.id
                AND data_encerramento IS NULL) contratos_ativos,
               (SELECT COUNT(*) FROM contrato WHERE estagiario_id=e.id) total_contratos
        FROM estagiario e ORDER BY e.nome""")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Estagiários'
    headers = ['ID', 'Nome', 'CPF', 'E-mail', 'Telefone', 'Cidade', 'Estado',
               'Curso', 'Semestre', 'Contratos Ativos', 'Total Contratos']
    _xlsx_apply_header(ws, headers)

    data_font = Font(name='Arial', size=10)
    alt_fill = PatternFill('solid', start_color='F2F6FC', end_color='F2F6FC')
    for i, r in enumerate(rows, 2):
        vals = [r['id'], r['nome'], r['cpf'] or '', r['email'] or '', r['telefone'] or '',
                r['cidade'] or '', r['estado'] or '', r['curso'] or '', r['semestre'] or '',
                r['contratos_ativos'], r['total_contratos']]
        for col, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=col, value=v)
            cell.font = data_font
            if i % 2 == 0:
                cell.fill = alt_fill

    col_widths = [6, 32, 16, 30, 16, 18, 8, 25, 10, 16, 14]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
    ws.freeze_panes = 'A2'
    return _xlsx_response(wb, f'estagiarios_{date.today()}.xlsx')


@app.route('/relatorios/indicadores.xlsx')
@login_required
def relatorio_indicadores_xlsx():
    hoje = date.today()
    d30 = hoje + timedelta(days=30)

    kpis = {
        'Contratos ativos': _q("SELECT COUNT(*) n FROM contrato WHERE data_encerramento IS NULL", one=True)['n'],
        'Contratos vencendo 30 dias': _q("SELECT COUNT(*) n FROM contrato WHERE data_encerramento IS NULL AND data_termino BETWEEN %s AND %s", (hoje, d30), one=True)['n'],
        'Contratos vencidos s/ TRE': _q("SELECT COUNT(*) n FROM contrato WHERE data_encerramento IS NULL AND data_termino < %s", (hoje,), one=True)['n'],
        'Estagiários cadastrados': _q("SELECT COUNT(*) n FROM estagiario", one=True)['n'],
        'Empresas cadastradas': _q("SELECT COUNT(*) n FROM empresa", one=True)['n'],
        'IEs cadastradas': _q("SELECT COUNT(*) n FROM ie", one=True)['n'],
        'Vagas abertas': _q("SELECT COUNT(*) n FROM vaga WHERE status='aberta'", one=True)['n'],
        'Leads CRM ativos': _q("SELECT COUNT(*) n FROM crm_lead WHERE status NOT IN ('perdido','cancelado')", one=True)['n'],
    }

    funil = _q("SELECT etapa, COUNT(*) qtd FROM crm_lead WHERE status NOT IN ('perdido','cancelado') GROUP BY etapa ORDER BY etapa")

    wb = openpyxl.Workbook()

    # Sheet 1 — KPIs
    ws1 = wb.active
    ws1.title = 'KPIs'
    _xlsx_apply_header(ws1, ['Indicador', 'Valor', 'Data de referência'])
    data_font = Font(name='Arial', size=10)
    bold_font = Font(name='Arial', size=10, bold=True)
    for i, (k, v) in enumerate(kpis.items(), 2):
        ws1.cell(row=i, column=1, value=k).font = data_font
        ws1.cell(row=i, column=2, value=v).font = bold_font
        ws1.cell(row=i, column=3, value=hoje.strftime('%d/%m/%Y')).font = data_font
    ws1.column_dimensions['A'].width = 35
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 20

    # Sheet 2 — Funil CRM
    ws2 = wb.create_sheet('Funil CRM')
    _xlsx_apply_header(ws2, ['Etapa', 'Quantidade'])
    for i, r in enumerate(funil, 2):
        ws2.cell(row=i, column=1, value=r['etapa']).font = data_font
        ws2.cell(row=i, column=2, value=r['qtd']).font = bold_font
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 12
    ws2.freeze_panes = 'A2'

    return _xlsx_response(wb, f'indicadores_{hoje}.xlsx')


# ─── PROSPECÇÃO COMERCIAL ────────────────────────────────────────────────────

@app.route('/crm/bairros')
@crm_required
def crm_bairros():
    cidade = request.args.get('cidade', '').strip()
    if not cidade:
        return jsonify([])
    db_bairros = [r['bairro'] for r in
                  _q("SELECT DISTINCT bairro FROM prospecto WHERE cidade=%s AND bairro IS NOT NULL AND bairro<>'' ORDER BY bairro", (cidade,))]
    seed = BAIRROS_VDC if cidade == 'Vitória da Conquista' else []
    merged = sorted(set(db_bairros) | set(seed))
    return jsonify(merged)


@app.route('/crm/prospeccao')
@crm_required
def crm_prospeccao():
    seg    = request.args.get('segmento', '')
    cnae   = request.args.get('cnae', '').strip()
    porte  = request.args.get('porte', '')
    cidade = request.args.get('cidade', '')
    bairro = request.args.get('bairro', '')
    status = request.args.get('status', '')

    sql = """SELECT p.*, u.nome as resp_nome FROM prospecto p
             LEFT JOIN usuario u ON u.id = p.responsavel_id
             WHERE 1=1"""
    params = []
    if seg:
        sql += " AND p.segmento=%s"; params.append(seg)
    if cnae:
        sql += " AND (p.cnae_codigo ILIKE %s OR p.cnae_descricao ILIKE %s)"; params += [f'%{cnae}%', f'%{cnae}%']
    if porte:
        sql += " AND p.porte=%s"; params.append(porte)
    if cidade:
        sql += " AND p.cidade=%s"; params.append(cidade)
    if bairro:
        sql += " AND p.bairro=%s"; params.append(bairro)
    if status:
        sql += " AND p.status=%s"; params.append(status)
    sql += " ORDER BY p.updated_at DESC"
    prospectos = _q(sql, params)

    cidades = [r['cidade'] for r in
               _q("SELECT DISTINCT cidade FROM prospecto WHERE cidade IS NOT NULL AND cidade<>'' ORDER BY cidade")]
    bairros_sel = []
    if cidade:
        bairros_sel = [r['bairro'] for r in
                       _q("SELECT DISTINCT bairro FROM prospecto WHERE cidade=%s AND bairro IS NOT NULL AND bairro<>'' ORDER BY bairro", (cidade,))]
        if cidade == 'Vitória da Conquista':
            bairros_sel = sorted(set(bairros_sel) | set(BAIRROS_VDC))

    por_status = {s: 0 for s in STATUS_PROSPECTO}
    por_seg    = {}
    for p in _q("SELECT segmento, status, COUNT(*) n FROM prospecto GROUP BY segmento, status"):
        por_status[p['status']] = por_status.get(p['status'], 0) + p['n']
        if p['segmento']:
            por_seg[p['segmento']] = por_seg.get(p['segmento'], 0) + p['n']

    return render_template('crm/prospeccao.html',
        prospectos=prospectos,
        segmentos=SEGMENTOS_PROSPECTO,
        portes=PORTES_EMPRESA,
        cnae_grupos=CNAE_GRUPOS,
        status_list=STATUS_PROSPECTO,
        cidades=cidades,
        bairros_sel=bairros_sel,
        por_status=por_status,
        por_seg=por_seg,
        seg_cor=SEGMENTO_COR,
        filtros=dict(segmento=seg, cnae=cnae, porte=porte, cidade=cidade, bairro=bairro, status=status),
    )


@app.route('/crm/prospeccao/novo', methods=['GET', 'POST'])
@crm_required
def crm_prospeccao_novo():
    if request.method == 'POST':
        _run("""INSERT INTO prospecto
                (empresa_nome,cnpj,segmento,cnae_codigo,cnae_descricao,porte,
                 cidade,bairro,endereco,telefone,email,contato_nome,contato_cargo,
                 site,vagas_estimadas,obs,status,responsavel_id)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)""",
             (request.form['empresa_nome'].strip(),
              request.form.get('cnpj') or None,
              request.form.get('segmento') or None,
              request.form.get('cnae_codigo') or None,
              request.form.get('cnae_descricao') or None,
              request.form.get('porte') or None,
              request.form.get('cidade') or 'Vitória da Conquista',
              request.form.get('bairro') or None,
              request.form.get('endereco') or None,
              request.form.get('telefone') or None,
              request.form.get('email') or None,
              request.form.get('contato_nome') or None,
              request.form.get('contato_cargo') or None,
              request.form.get('site') or None,
              request.form.get('vagas_estimadas') or None,
              request.form.get('obs') or None,
              request.form.get('status', 'novo'),
              request.form.get('responsavel_id') or current_user.id))
        flash('Prospecto cadastrado!', 'success')
        return redirect(url_for('crm_prospeccao'))
    return render_template('crm/prospeccao_form.html',
        p=None, segmentos=SEGMENTOS_PROSPECTO, portes=PORTES_EMPRESA,
        cnae_grupos=CNAE_GRUPOS, status_list=STATUS_PROSPECTO,
        bairros_vdc=BAIRROS_VDC, usuarios_crm=_crm_usuarios(),
        pode_ver_todos=_crm_pode_ver_todos(), current_user_id=int(current_user.id))


@app.route('/crm/prospeccao/<int:id>/editar', methods=['GET', 'POST'])
@crm_required
def crm_prospeccao_editar(id):
    p = _q("SELECT * FROM prospecto WHERE id=%s", (id,), one=True)
    if not p:
        abort(404)
    if request.method == 'POST':
        _run("""UPDATE prospecto SET
                empresa_nome=%s,cnpj=%s,segmento=%s,cnae_codigo=%s,cnae_descricao=%s,
                porte=%s,cidade=%s,bairro=%s,endereco=%s,telefone=%s,email=%s,
                contato_nome=%s,contato_cargo=%s,site=%s,vagas_estimadas=%s,obs=%s,
                status=%s,responsavel_id=%s,updated_at=CURRENT_TIMESTAMP
                WHERE id=%s""",
             (request.form['empresa_nome'].strip(),
              request.form.get('cnpj') or None,
              request.form.get('segmento') or None,
              request.form.get('cnae_codigo') or None,
              request.form.get('cnae_descricao') or None,
              request.form.get('porte') or None,
              request.form.get('cidade') or 'Vitória da Conquista',
              request.form.get('bairro') or None,
              request.form.get('endereco') or None,
              request.form.get('telefone') or None,
              request.form.get('email') or None,
              request.form.get('contato_nome') or None,
              request.form.get('contato_cargo') or None,
              request.form.get('site') or None,
              request.form.get('vagas_estimadas') or None,
              request.form.get('obs') or None,
              request.form.get('status', 'novo'),
              request.form.get('responsavel_id') or current_user.id,
              id))
        flash('Prospecto atualizado!', 'success')
        return redirect(url_for('crm_prospeccao'))
    bairros_cidade = BAIRROS_VDC if (p['cidade'] or '') == 'Vitória da Conquista' else []
    return render_template('crm/prospeccao_form.html',
        p=p, segmentos=SEGMENTOS_PROSPECTO, portes=PORTES_EMPRESA,
        cnae_grupos=CNAE_GRUPOS, status_list=STATUS_PROSPECTO,
        bairros_vdc=bairros_cidade, usuarios_crm=_crm_usuarios(),
        pode_ver_todos=_crm_pode_ver_todos(), current_user_id=int(current_user.id))


@app.route('/crm/prospeccao/<int:id>/excluir', methods=['POST'])
@crm_required
def crm_prospeccao_excluir(id):
    _run("DELETE FROM prospecto WHERE id=%s", (id,))
    flash('Prospecto excluído.', 'success')
    return redirect(url_for('crm_prospeccao'))


@app.route('/crm/prospeccao/<int:id>/converter', methods=['POST'])
@crm_required
def crm_prospeccao_converter(id):
    p = _q("SELECT * FROM prospecto WHERE id=%s", (id,), one=True)
    if not p:
        abort(404)
    lead_id = _ins("""INSERT INTO crm_lead
        (empresa_nome,empresa_cnpj,cidade,segmento,vagas_estimadas,etapa,
         origem,responsavel_id,contato_nome,contato_email,contato_whatsapp,obs)
        VALUES (%s,%s,%s,%s,%s,'Lead Captado','Prospecção ativa',%s,%s,%s,%s,%s)""",
        (p['empresa_nome'], p['cnpj'], p['cidade'], p['segmento'],
         p['vagas_estimadas'], p['responsavel_id'] or current_user.id,
         p['contato_nome'], p['email'], p['telefone'], p['obs']))
    _run("UPDATE prospecto SET status='convertido', lead_id=%s, updated_at=CURRENT_TIMESTAMP WHERE id=%s",
         (lead_id, id))
    flash(f'"{p["empresa_nome"]}" convertido em lead no CRM!', 'success')
    return redirect(url_for('crm_lead_detalhe', id=lead_id))


init_db()


try:
    from apscheduler.schedulers.background import BackgroundScheduler
    _scheduler = BackgroundScheduler(timezone='America/Bahia')
    _scheduler.add_job(enviar_backup_email, 'cron', hour=2, minute=0)
    _scheduler.start()
except Exception as _e:
    print(f'[Scheduler] Não iniciado: {_e}')

if __name__ == '__main__':
    print('\n' + '=' * 50)
    print('  CICLO RH — Sistema de Estágio')
    print('  Acesse: http://localhost:5000')
    print('=' * 50 + '\n')
    app.run(debug=False, port=5000)
