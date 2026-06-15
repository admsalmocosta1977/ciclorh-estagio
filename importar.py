"""
importar.py — Importa dados de uma planilha .xlsx/.xlsm para o PostgreSQL.

Uso via linha de comando (apontando para o banco remoto):
    DATABASE_URL=postgresql://... python importar.py

Uso via web (chamado pelo Flask):
    resultado = run_from_file(caminho_arquivo, database_url)
"""

import os, sys
from datetime import datetime

try:
    import openpyxl
except ImportError:
    sys.exit("Instale openpyxl:  pip install openpyxl")

try:
    import psycopg2, psycopg2.extras
except ImportError:
    sys.exit("Instale psycopg2:  pip install psycopg2-binary")

# Mapa de colunas (letra A=0, B=1, …) — ajuste se necessário
COL = {
    'nome':        0,   # A — Nome do estagiário
    'cpf':         1,   # B
    'empresa':     2,   # C — Nome da empresa/concedente
    'ie':          3,   # D — Instituição de ensino
    'curso':       4,   # E
    'data_inicio': 5,   # F
    'data_fim':    6,   # G
    'bolsa':       7,   # H
    'taxa':        8,   # I
    'supervisor':  9,   # J
    'coordenador': 10,  # K
    'area':        11,  # L
    'tipo':        12,  # M
    'ch_diaria':   13,  # N
    'ch_semanal':  14,  # O
}


def _clean(val):
    if val is None:
        return None
    s = str(val).strip()
    return s if s and s.lower() not in ('#n/a', 'n/a', '#ref!', '#value!', '') else None


def _date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d-%m-%Y', '%d/%m/%y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except Exception:
            pass
    return None


def _float(val):
    try:
        return float(str(val).replace(',', '.').replace('R$', '').strip())
    except Exception:
        return None


def _cell(row, col_idx):
    return _clean(row[col_idx]) if col_idx < len(row) else None


def run_from_file(filepath, database_url):
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)

    # Usa aba 'DADOS' se existir, senão a primeira aba
    sheet_name = 'DADOS' if 'DADOS' in wb.sheetnames else wb.sheetnames[0]
    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not all_rows:
        return dict(ok=0, skip=0, empresas=0, ies=0, estagiarios=0, erro='Aba vazia')

    # Detecta linha de cabeçalho (primeira linha não vazia)
    header_idx = next((i for i, r in enumerate(all_rows) if any(c is not None for c in r)), 0)
    data_rows = all_rows[header_idx + 1:]

    url = database_url
    if url.startswith('postgres://'):
        url = url.replace('postgres://', 'postgresql://', 1)

    conn = psycopg2.connect(url, cursor_factory=psycopg2.extras.RealDictCursor)
    conn.autocommit = True
    cur = conn.cursor()

    empresas_map, ies_map, estagiarios_map = {}, {}, {}
    ok = skip = 0

    for row in data_rows:
        nome      = _cell(row, COL['nome'])
        cpf_raw   = _cell(row, COL['cpf'])
        empresa   = _cell(row, COL['empresa'])
        ie_nome   = _cell(row, COL['ie'])
        curso     = _cell(row, COL['curso'])
        d_ini     = _date(_cell(row, COL['data_inicio']))
        d_fim     = _date(_cell(row, COL['data_fim']))

        if not nome or not cpf_raw or not empresa or not d_ini or not d_fim:
            skip += 1
            continue

        cpf = ''.join(c for c in cpf_raw if c.isdigit())
        if len(cpf) != 11:
            cpf = cpf_raw

        # Empresa
        if empresa not in empresas_map:
            cur.execute("SELECT id FROM empresa WHERE nome = %s", (empresa,))
            r = cur.fetchone()
            if r:
                empresas_map[empresa] = r['id']
            else:
                sup = _cell(row, COL['supervisor'])
                cur.execute("INSERT INTO empresa (nome, supervisor_nome) VALUES (%s, %s) RETURNING id",
                            (empresa, sup))
                empresas_map[empresa] = cur.fetchone()['id']

        # IE
        ie_key = ie_nome or 'NÃO INFORMADA'
        if ie_key not in ies_map:
            cur.execute("SELECT id FROM ie WHERE nome = %s", (ie_key,))
            r = cur.fetchone()
            if r:
                ies_map[ie_key] = r['id']
            else:
                coord = _cell(row, COL['coordenador'])
                cur.execute("INSERT INTO ie (nome, coordenador) VALUES (%s, %s) RETURNING id",
                            (ie_key, coord))
                ies_map[ie_key] = cur.fetchone()['id']

        # Estagiário
        if cpf not in estagiarios_map:
            cur.execute("SELECT id FROM estagiario WHERE cpf = %s", (cpf,))
            r = cur.fetchone()
            if r:
                estagiarios_map[cpf] = r['id']
            else:
                try:
                    cur.execute("INSERT INTO estagiario (nome, cpf) VALUES (%s, %s) RETURNING id",
                                (nome, cpf))
                    estagiarios_map[cpf] = cur.fetchone()['id']
                except psycopg2.errors.UniqueViolation:
                    conn.autocommit = False
                    conn.rollback()
                    conn.autocommit = True
                    cur.execute("SELECT id FROM estagiario WHERE cpf = %s", (cpf,))
                    estagiarios_map[cpf] = cur.fetchone()['id']

        # Contrato
        bolsa    = _float(_cell(row, COL['bolsa']))
        taxa     = _float(_cell(row, COL['taxa']))
        area     = _cell(row, COL['area'])
        tipo     = _cell(row, COL['tipo']) or 'Não Obrigatório'
        sup_nome = _cell(row, COL['supervisor'])
        try:
            ch_d = int(float(_cell(row, COL['ch_diaria']) or 6))
        except Exception:
            ch_d = 6
        try:
            ch_s = int(float(_cell(row, COL['ch_semanal']) or 30))
        except Exception:
            ch_s = 30

        cur.execute("""
            INSERT INTO contrato
              (estagiario_id, empresa_id, ie_id, supervisor_nome, curso,
               tipo_estagio, area_atuacao, ch_diaria, ch_semanal,
               data_inicio, data_fim, bolsa, taxa)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT DO NOTHING
        """, (estagiarios_map[cpf], empresas_map[empresa], ies_map[ie_key],
              sup_nome, curso or 'Não informado', tipo, area,
              ch_d, ch_s, d_ini, d_fim, bolsa, taxa))
        ok += 1

    cur.close()
    conn.close()

    return dict(ok=ok, skip=skip,
                empresas=len(empresas_map),
                ies=len(ies_map),
                estagiarios=len(estagiarios_map))


if __name__ == '__main__':
    db_url = os.environ.get('DATABASE_URL', '')
    if not db_url:
        sys.exit('Defina DATABASE_URL antes de executar:\n'
                 '  $env:DATABASE_URL="postgresql://..."\n'
                 '  python importar.py')

    planilha = os.environ.get('PLANILHA',
               r'C:\Users\usuario\OneDrive\Área de Trabalho\PLANILHA SALMO.xlsm')
    if not os.path.exists(planilha):
        sys.exit(f'Planilha não encontrada: {planilha}')

    print(f'Importando: {planilha}')
    r = run_from_file(planilha, db_url)
    print(f'\n{"="*40}')
    print(f'  Contratos importados : {r["ok"]}')
    print(f'  Linhas ignoradas     : {r["skip"]}')
    print(f'  Empresas             : {r["empresas"]}')
    print(f'  Instituições         : {r["ies"]}')
    print(f'  Estagiários          : {r["estagiarios"]}')
    print(f'{"="*40}')
